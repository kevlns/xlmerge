"""Three-way, sheet-aware XLSX conflict detection and merge primitives."""

from __future__ import annotations

import hashlib
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

# 兼容嵌入式 Python（._pth 忽略 PYTHONPATH）：包内自举，将包根加入 sys.path
_PKG_ROOT = Path(__file__).resolve().parents[1]
if str(_PKG_ROOT) not in sys.path:
    sys.path.insert(0, str(_PKG_ROOT))

from xlsx_merge_engine.xlsx_git_merge_bridge import analyze_workbooks as analyze_structural_workbooks


class ConflictError(RuntimeError):
    """Raised when a conflict cannot be prepared or resolved safely."""


WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm"}


def _is_supported_workbook(path: str | Path) -> bool:
    return Path(path).suffix.lower() in WORKBOOK_EXTENSIONS


def _is_macro_workbook(path: str | Path) -> bool:
    return Path(path).suffix.lower() == ".xlsm"


def _is_macro_payload_entry(name: str) -> bool:
    lowered = name.lower()
    return (
        lowered.startswith(("xl/vba", "xl/ctrlprops/", "customui/", "xl/activex/"))
        or (lowered.startswith("xl/drawings/") and "vmldrawing" in lowered and lowered.endswith(".vml"))
        or (lowered.startswith("xl/media/") and lowered.endswith(".emf"))
    )


def _macro_payload_digest(source: bytes | Path | str) -> tuple[str, list[str]]:
    raw = source if isinstance(source, bytes) else Path(source).read_bytes()
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            entries = sorted(name for name in archive.namelist() if _is_macro_payload_entry(name))
            digest = hashlib.sha256()
            for name in entries:
                digest.update(name.encode("utf-8"))
                digest.update(b"\0")
                digest.update(archive.read(name))
            return digest.hexdigest(), entries
    except (OSError, zipfile.BadZipFile) as exc:
        raise ConflictError(f"Invalid macro-enabled workbook package: {exc}") from exc


def _macro_payload_info(base_blob: bytes | None, ours_blob: bytes, theirs_blob: bytes) -> dict[str, Any]:
    base_hash, base_entries = _macro_payload_digest(base_blob) if base_blob is not None else (None, [])
    ours_hash, ours_entries = _macro_payload_digest(ours_blob)
    theirs_hash, theirs_entries = _macro_payload_digest(theirs_blob)
    ours_key = (ours_hash, ours_entries)
    theirs_key = (theirs_hash, theirs_entries)
    base_key = (base_hash, base_entries) if base_hash is not None else None
    compatible = ours_entries == theirs_entries
    if ours_key == theirs_key:
        status, default = "same", "ours"
    elif not compatible:
        status, default = "incompatible", "ours"
    elif base_key is not None and ours_key == base_key:
        status, default = "auto_theirs", "theirs"
    elif base_key is not None and theirs_key == base_key:
        status, default = "auto_ours", "ours"
    else:
        status, default = "conflict", "ours"
    return {
        "status": status,
        "default": default,
        "canChooseTheirs": compatible,
        "baseHash": base_hash,
        "oursHash": ours_hash,
        "theirsHash": theirs_hash,
        "baseEntries": base_entries,
        "oursEntries": ours_entries,
        "theirsEntries": theirs_entries,
    }


def _replace_macro_payload(target: Path, source: Path | str) -> None:
    source_path = Path(source)
    with zipfile.ZipFile(target) as target_archive, zipfile.ZipFile(source_path) as source_archive:
        target_parts = [
            (info, target_archive.read(info.filename))
            for info in target_archive.infolist()
            if not _is_macro_payload_entry(info.filename)
        ]
        source_parts = [
            (info, source_archive.read(info.filename))
            for info in source_archive.infolist()
            if _is_macro_payload_entry(info.filename)
        ]
    fd, replacement_name = tempfile.mkstemp(prefix=f".{target.stem}-macro-", suffix=target.suffix, dir=target.parent)
    os.close(fd)
    replacement = Path(replacement_name)
    try:
        with zipfile.ZipFile(replacement, "w") as output_archive:
            for info, payload in [*target_parts, *source_parts]:
                output_archive.writestr(info, payload)
        os.replace(replacement, target)
    finally:
        replacement.unlink(missing_ok=True)


def _close_workbook(workbook) -> None:
    if workbook is None:
        return
    vba_archive = getattr(workbook, "vba_archive", None)
    workbook.close()
    if vba_archive is not None:
        vba_archive.close()


def _run_git(repo: Path, args: list[str], *, binary: bool = False, check: bool = False):
    result = subprocess.run(
        ["git", "-C", str(repo), "-c", "core.quotepath=false", *args],
        capture_output=True,
        text=not binary,
        encoding=None if binary else "utf-8",
        errors=None if binary else "replace",
        timeout=60,
    )
    if check and result.returncode != 0:
        stderr = result.stderr.decode("utf-8", "replace") if binary else result.stderr
        raise ConflictError(f"git {' '.join(args)} failed: {stderr.strip()}")
    return result


def find_repo_root(repo: Path | str) -> Path:
    candidate = Path(repo).resolve()
    result = _run_git(candidate, ["rev-parse", "--show-toplevel"])
    if result.returncode != 0:
        raise ConflictError(f"Not a Git repository: {candidate}")
    return Path(result.stdout.strip()).resolve()


def detect_conflicts(repo: Path | str) -> list[dict[str, Any]]:
    repo_root = find_repo_root(repo)
    result = _run_git(repo_root, ["diff", "--name-only", "--diff-filter=U", "-z"], binary=True, check=True)
    paths = result.stdout.decode("utf-8", "surrogateescape").split("\0")
    conflicts: list[dict[str, Any]] = []
    for relative_path in paths:
        if not relative_path or not _is_supported_workbook(relative_path):
            continue
        stages_result = _run_git(repo_root, ["ls-files", "-u", "--", relative_path], check=True)
        stages: list[int] = []
        for line in stages_result.stdout.splitlines():
            prefix = line.split("\t", 1)[0].split()
            if len(prefix) >= 3:
                stages.append(int(prefix[2]))
        conflicts.append({"path": relative_path.replace("\\", "/"), "stages": sorted(set(stages))})
    return conflicts


def _git_blob(repo: Path, stage: int, relative_path: str) -> bytes | None:
    result = _run_git(repo, ["show", f":{stage}:{relative_path}"], binary=True)
    return result.stdout if result.returncode == 0 else None


def _safe_runtime_dir(repo: Path, runtime_dir: Path | str | None) -> Path:
    if runtime_dir:
        target = Path(runtime_dir).resolve()
    else:
        repo_key = str(abs(hash(str(repo).lower())))
        target = Path(tempfile.gettempdir()) / "xlmerge" / repo_key
    target.mkdir(parents=True, exist_ok=True)
    return target


def _slug(relative_path: str) -> str:
    safe = "".join(ch if ch.isalnum() or ch in "-_." else "_" for ch in relative_path)
    return safe[-120:] or "workbook"


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _cell_token(cell, cached_cell=None) -> tuple[Any, str] | None:
    if cell is None or isinstance(cell, MergedCell):
        return None
    # 公式 cell 按缓存值参与比较，避免公式不变但引用值变了被误判为 same
    if cell.data_type == 'f' and cached_cell is not None and not isinstance(cached_cell, MergedCell):
        return cached_cell.value, cached_cell.data_type
    return cell.value, cell.data_type


def _cell_display_value(cell, cached_cell) -> Any:
    """UI 展示值：公式 cell 取缓存值，其余取原始值。"""
    if cell is None or isinstance(cell, MergedCell):
        return None
    if cell.data_type == 'f' and cached_cell is not None and not isinstance(cached_cell, MergedCell):
        return cached_cell.value
    return cell.value


def _value_cells(ws) -> dict[tuple[int, int], Any]:
    if ws is None:
        return {}
    result = {}
    for key, cell in ws._cells.items():
        if isinstance(cell, MergedCell):
            continue
        if cell.value is not None:
            result[key] = cell
    return result


def _sheet_values(ws) -> dict[tuple[int, int], tuple[Any, str] | None]:
    return {key: _cell_token(cell) for key, cell in _value_cells(ws).items()}


def _sheet_equal(left, right) -> bool:
    return _sheet_values(left) == _sheet_values(right)


def _ordered_sheet_names(*workbooks) -> list[str]:
    names: list[str] = []
    for workbook in workbooks:
        if workbook is None:
            continue
        for name in workbook.sheetnames:
            if name not in names:
                names.append(name)
    return names


def _context_rows(ws, ws_v, rows: set[int], max_col: int) -> dict[str, list[Any]]:
    if ws is None or max_col <= 0:
        return {}
    header_rows: set[int] = set()
    for row in range(1, min(max(ws.max_row or 1, 1), 20) + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.startswith("##"):
            header_rows.add(row)
        elif header_rows:
            break
    wanted = sorted(rows | header_rows)
    return {
        str(row): [_json_value(_cell_display_value(
            ws.cell(row, col),
            ws_v.cell(row, col) if ws_v is not None else None
        )) for col in range(1, max_col + 1)]
        for row in wanted
    }


def _modified_sheet_diff(name, ws_base, ws_ours, ws_theirs,
                         ws_base_v=None, ws_ours_v=None, ws_theirs_v=None) -> dict[str, Any]:
    base_cells = _value_cells(ws_base)
    ours_cells = _value_cells(ws_ours)
    theirs_cells = _value_cells(ws_theirs)
    coordinates = sorted(set(base_cells) | set(ours_cells) | set(theirs_cells))
    cells: list[dict[str, Any]] = []
    changed_rows: set[int] = set()

    for row, col in coordinates:
        base_cell = base_cells.get((row, col))
        ours_cell = ours_cells.get((row, col))
        theirs_cell = theirs_cells.get((row, col))
        base_cached = ws_base_v.cell(row, col) if ws_base_v is not None else None
        ours_cached = ws_ours_v.cell(row, col) if ws_ours_v is not None else None
        theirs_cached = ws_theirs_v.cell(row, col) if ws_theirs_v is not None else None
        base_token = _cell_token(base_cell, base_cached)
        ours_token = _cell_token(ours_cell, ours_cached)
        theirs_token = _cell_token(theirs_cell, theirs_cached)
        if ours_token == theirs_token:
            continue
        if ours_token == base_token:
            resolution = "theirs"
        elif theirs_token == base_token:
            resolution = "ours"
        else:
            resolution = "conflict"
        changed_rows.add(row)
        cells.append({
            "row": row,
            "col": col,
            "coordinate": f"{get_column_letter(col)}{row}",
            "base": _json_value(_cell_display_value(base_cell, base_cached)),
            "ours": _json_value(_cell_display_value(ours_cell, ours_cached)),
            "theirs": _json_value(_cell_display_value(theirs_cell, theirs_cached)),
            "resolution": resolution,
        })

    preferred_ws = ws_ours or ws_theirs or ws_base
    preferred_ws_v = ws_ours_v or ws_theirs_v or ws_base_v
    max_col = max([col for _, col in coordinates] or [1])
    return {
        "name": name,
        "action": "merge",
        "cells": cells,
        "contextRows": _context_rows(preferred_ws, preferred_ws_v, changed_rows, max_col),
        "maxCol": max_col,
        "changedCount": len(cells),
        "conflictCount": sum(cell["resolution"] == "conflict" for cell in cells),
    }


def _axis_state(status: str, *, ours_present: bool) -> dict[str, Any]:
    if status.startswith("added_"):
        return {
            "state": "added",
            "locked": True,
            "incoming": status == "added_theirs",
            "default": "keep",
        }
    if status.startswith("delete_modify_"):
        return {
            "state": "delete_modify",
            "locked": False,
            "incoming": False,
            "default": "keep" if ours_present else "delete",
        }
    if status.startswith("deleted_"):
        return {
            "state": "deleted",
            "locked": True,
            "incoming": False,
            "default": "delete",
        }
    return {"state": "common", "locked": False, "incoming": False, "default": "keep"}


def _side_cell(workbook, sheet_name: str | None, row: int | None, column: int | None):
    if workbook is None or sheet_name is None or row is None or column is None:
        return None
    return workbook[sheet_name].cell(row, column)


def _cell_resolution(base_cell, ours_cell, theirs_cell, *, has_base: bool,
                     base_cached=None, ours_cached=None, theirs_cached=None) -> str:
    base_token = _cell_token(base_cell, base_cached)
    ours_token = _cell_token(ours_cell, ours_cached)
    theirs_token = _cell_token(theirs_cell, theirs_cached)
    if not has_base:
        return "theirs" if theirs_cell is not None else "ours"
    if ours_token == theirs_token:
        return "same"
    if ours_token == base_token:
        return "theirs"
    if theirs_token == base_token:
        return "ours"
    return "conflict"


def _header_rows(ws) -> list[int]:
    if ws is None:
        return []
    rows = []
    for row in range(1, min(ws.max_row or 0, 40) + 1):
        marker = ws.cell(row, 1).value
        if isinstance(marker, str) and marker.startswith("##"):
            rows.append(row)
        elif rows:
            break
    return rows


def _merged_region_key(region: dict[str, Any]) -> tuple[tuple[str, ...], tuple[str, ...]]:
    return tuple(region["rows"]), tuple(region["columns"])


def _canonical_merged_regions(ws, side: str, rows, columns) -> list[dict[str, Any]]:
    if ws is None:
        return []
    header_rows = set(_header_rows(ws))
    row_lookup = {
        row.get(f"{side}Row"): row["id"]
        for row in rows
        if row.get(f"{side}Row") is not None
    }
    column_lookup = {
        column.get(f"{side}Col"): column["id"]
        for column in columns
        if column.get(f"{side}Col") is not None
    }
    regions = []
    for merged_range in ws.merged_cells.ranges:
        row_ids = []
        for row_number in range(merged_range.min_row, merged_range.max_row + 1):
            row_id = f"header:{row_number}" if row_number in header_rows else row_lookup.get(row_number)
            if row_id is None:
                break
            row_ids.append(row_id)
        else:
            column_ids = []
            for column_number in range(merged_range.min_col, merged_range.max_col + 1):
                column_id = column_lookup.get(column_number)
                if column_id is None:
                    break
                column_ids.append(column_id)
            else:
                regions.append({"rows": row_ids, "columns": column_ids})
    return sorted(regions, key=_merged_region_key)


def _regions_overlap(regions: list[dict[str, Any]]) -> bool:
    occupied: set[tuple[str, str]] = set()
    for region in regions:
        cells = {(row, column) for row in region["rows"] for column in region["columns"]}
        if occupied & cells:
            return True
        occupied.update(cells)
    return False


def _merged_cells_diff(names, rows, columns, wb_base, wb_ours, wb_theirs) -> dict[str, Any]:
    workbooks = {"base": wb_base, "ours": wb_ours, "theirs": wb_theirs}
    regions = {}
    for side, workbook in workbooks.items():
        sheet_name = names.get(side)
        ws = workbook[sheet_name] if workbook is not None and sheet_name in workbook.sheetnames else None
        regions[side] = _canonical_merged_regions(ws, side, rows, columns)

    keyed = {side: {_merged_region_key(item): item for item in items} for side, items in regions.items()}
    base_keys, ours_keys, theirs_keys = (set(keyed[side]) for side in ("base", "ours", "theirs"))
    if ours_keys == theirs_keys:
        status, resolved = "same", regions["ours"]
    elif ours_keys == base_keys:
        status, resolved = "auto_theirs", regions["theirs"]
    elif theirs_keys == base_keys:
        status, resolved = "auto_ours", regions["ours"]
    else:
        all_keys = base_keys | ours_keys | theirs_keys
        resolved_keys = set()
        for key in all_keys:
            base_present = key in base_keys
            ours_present = key in ours_keys
            theirs_present = key in theirs_keys
            if ours_present == theirs_present:
                keep = ours_present
            elif ours_present == base_present:
                keep = theirs_present
            else:
                keep = ours_present
            if keep:
                resolved_keys.add(key)
        source_items = {**keyed["base"], **keyed["ours"], **keyed["theirs"]}
        candidate = [source_items[key] for key in sorted(resolved_keys)]
        if _regions_overlap(candidate):
            status, resolved = "conflict", []
        else:
            status, resolved = "auto_merge", candidate
    return {
        "status": status,
        "default": "ours" if status == "conflict" else None,
        "base": regions["base"],
        "ours": regions["ours"],
        "theirs": regions["theirs"],
        "resolved": resolved,
    }


def _structural_sheet_diff(report_sheet, wb_base, wb_ours, wb_theirs,
                           wb_base_v=None, wb_ours_v=None, wb_theirs_v=None) -> dict[str, Any]:
    names = report_sheet.get("names") or {
        "base": report_sheet["name"], "ours": report_sheet["name"], "theirs": report_sheet["name"]
    }
    rows = []
    for index, entry in enumerate(report_sheet.get("rowAlignment") or [], 1):
        item = dict(entry)
        item.update(_axis_state(entry["status"], ours_present=entry.get("oursRow") is not None))
        item["display"] = entry.get("oursRow") or entry.get("baseRow") or entry.get("theirsRow") or index
        rows.append(item)
    columns = []
    for index, entry in enumerate(report_sheet.get("columnAlignment") or [], 1):
        item = dict(entry)
        item["id"] = entry["field"]
        item.update(_axis_state(entry["status"], ours_present=entry.get("oursColumn") is not None))
        item["baseCol"] = column_index_from_string(entry["baseColumn"]) if entry.get("baseColumn") else None
        item["oursCol"] = column_index_from_string(entry["oursColumn"]) if entry.get("oursColumn") else None
        item["theirsCol"] = column_index_from_string(entry["theirsColumn"]) if entry.get("theirsColumn") else None
        item["display"] = entry.get("oursColumn") or entry.get("baseColumn") or entry.get("theirsColumn") or get_column_letter(index)
        columns.append(item)

    all_cells: list[dict[str, Any]] = []
    changed_row_ids: set[str] = set()
    row_attention: set[str] = set()
    for row in rows:
        for column in columns:
            base_cell = _side_cell(wb_base, names.get("base"), row.get("baseRow"), column.get("baseCol"))
            ours_cell = _side_cell(wb_ours, names.get("ours"), row.get("oursRow"), column.get("oursCol"))
            theirs_cell = _side_cell(wb_theirs, names.get("theirs"), row.get("theirsRow"), column.get("theirsCol"))
            base_cached = _side_cell(wb_base_v, names.get("base"), row.get("baseRow"), column.get("baseCol")) if wb_base_v else None
            ours_cached = _side_cell(wb_ours_v, names.get("ours"), row.get("oursRow"), column.get("oursCol")) if wb_ours_v else None
            theirs_cached = _side_cell(wb_theirs_v, names.get("theirs"), row.get("theirsRow"), column.get("theirsCol")) if wb_theirs_v else None
            base_token = _cell_token(base_cell, base_cached)
            ours_token = _cell_token(ours_cell, ours_cached)
            theirs_token = _cell_token(theirs_cell, theirs_cached)
            has_base = row.get("baseRow") is not None and column.get("baseCol") is not None
            changed = not (ours_token == theirs_token == base_token)
            structural = row["state"] != "common" or column["state"] != "common"
            selectable = (
                row["state"] in ("common", "delete_modify")
                and column["state"] in ("common", "delete_modify")
            )
            if changed or structural:
                changed_row_ids.add(row["id"])
            resolution = _cell_resolution(base_cell, ours_cell, theirs_cell, has_base=has_base,
                                          base_cached=base_cached, ours_cached=ours_cached, theirs_cached=theirs_cached)
            if changed and resolution != "same":
                row_attention.add(row["id"])
            stable_id = f'{row["id"]}|{column["id"]}'
            display_coordinate = (
                f'{get_column_letter(column["oursCol"])}{row["oursRow"]}'
                if row.get("oursRow") is not None and column.get("oursCol") is not None
                else stable_id
            )
            all_cells.append({
                "id": stable_id,
                "coordinate": display_coordinate,
                "rowId": row["id"],
                "colId": column["id"],
                "base": _json_value(_cell_display_value(base_cell, base_cached)),
                "ours": _json_value(_cell_display_value(ours_cell, ours_cached)),
                "theirs": _json_value(_cell_display_value(theirs_cell, theirs_cached)),
                "basePresent": base_cell is not None,
                "oursPresent": ours_cell is not None,
                "theirsPresent": theirs_cell is not None,
                "resolution": resolution,
                "changed": changed,
                "structural": structural,
                "selectable": selectable,
            })

    # 只显示需要关注的行：结构变化行 + 含非 same（需决策/有变化信息）cell 的行
    visible_rows = [
        row for row in rows
        if row["state"] != "common" or row["id"] in row_attention
    ]
    visible_ids = {row["id"] for row in visible_rows}
    grid_cells = [cell for cell in all_cells if cell["rowId"] in visible_ids]
    cells = [cell for cell in grid_cells if cell["changed"]]
    decision_cell_count = sum(cell["changed"] and cell["selectable"] for cell in cells)
    preferred_ws = None
    for workbook, key in ((wb_ours, "ours"), (wb_theirs, "theirs"), (wb_base, "base")):
        if workbook is not None and names.get(key) in workbook.sheetnames:
            preferred_ws = workbook[names[key]]
            break
    header_rows = []
    for row_number in _header_rows(preferred_ws):
        header_rows.append({
            "row": row_number,
            "cells": [
                _json_value(preferred_ws.cell(row_number, column.get("oursCol") or column.get("theirsCol") or column.get("baseCol")).value)
                for column in columns
            ],
        })
    merged_cells = _merged_cells_diff(names, rows, columns, wb_base, wb_ours, wb_theirs)
    merged_cells_conflict = merged_cells["status"] == "conflict"
    structural_conflicts = (
        sum(axis["state"] == "delete_modify" for axis in [*rows, *columns])
        + int(merged_cells_conflict)
    )
    rebuild_required = (
        any(axis["state"] != "common" for axis in [*rows, *columns])
        or any(cell["changed"] for cell in all_cells)
        or merged_cells["status"] != "same"
    )
    selectable_cells = [
        cell for cell in cells
        if cell["changed"] and cell["resolution"] == "conflict"
        and next(row for row in rows if row["id"] == cell["rowId"])["state"] == "common"
        and next(column for column in columns if column["id"] == cell["colId"])["state"] == "common"
    ]
    return {
        "name": report_sheet["name"],
        "names": names,
        "renameStatus": report_sheet.get("renameStatus", "unchanged"),
        "action": "merge",
        "rows": visible_rows,
        "allRows": rows,
        "columns": columns,
        "cells": cells,
        "gridCells": grid_cells,
        "allCells": all_cells,
        "headerRows": header_rows,
        "mergedCells": merged_cells,
        "rebuildRequired": rebuild_required,
        "maxCol": len(columns),
        "changedCount": decision_cell_count,
        "decisionCellCount": decision_cell_count,
        "rawChangedCellCount": sum(cell["changed"] for cell in cells),
        "conflictCount": len(selectable_cells),
        "structuralConflictCount": structural_conflicts,
        "needsAttention": bool(selectable_cells) or structural_conflicts > 0
                         or report_sheet.get("renameStatus") == "rename_conflict",
    }


def diff_workbooks(
    base_path: Path | str | None,
    ours_path: Path | str,
    theirs_path: Path | str,
    *,
    analysis_dir: Path | str | None = None,
    skip_hidden_sheets: bool = True,
) -> dict[str, Any]:
    wb_base = load_workbook(base_path, data_only=False, keep_links=True) if base_path else None
    wb_ours = load_workbook(ours_path, data_only=False, keep_links=True)
    wb_theirs = load_workbook(theirs_path, data_only=False, keep_links=True)
    wb_base_v = load_workbook(base_path, data_only=True, keep_links=True) if base_path else None
    wb_ours_v = load_workbook(ours_path, data_only=True, keep_links=True)
    wb_theirs_v = load_workbook(theirs_path, data_only=True, keep_links=True)
    try:
        if base_path is not None:
            structural_dir = Path(analysis_dir) if analysis_dir else Path(tempfile.mkdtemp(prefix="xlsx-merge-analysis-"))
            structural = analyze_structural_workbooks(
                base_path, ours_path, theirs_path, structural_dir,
                write_sidecars=False, skip_hidden_sheets=skip_hidden_sheets,
            )
            sheets = []
            for report_sheet in structural["sheets"]:
                if report_sheet["action"] == "merge":
                    sheets.append(_structural_sheet_diff(
                        report_sheet, wb_base, wb_ours, wb_theirs,
                        wb_base_v, wb_ours_v, wb_theirs_v,
                    ))
                    continue
                names = report_sheet.get("names") or {}
                if not names.get("base"):
                    if names.get("ours") and names.get("theirs"):
                        action = "auto_add_theirs"
                    else:
                        action = "auto_add_ours" if names.get("ours") else "auto_add_theirs"
                elif report_sheet.get("status") == "conflict":
                    conflict_kind = (report_sheet.get("sheetConflict") or {}).get("kind", "")
                    if conflict_kind.startswith("delete_modify_"):
                        action = f'conflict_delete_{conflict_kind.rsplit("_", 1)[-1]}'
                    else:
                        action = "conflict_sheet_structure"
                else:
                    action = "auto_delete"
                sheets.append({
                    "name": report_sheet["name"],
                    "names": names,
                    "renameStatus": report_sheet.get("renameStatus", "unchanged"),
                    "action": action,
                    "sheetConflict": report_sheet.get("sheetConflict"),
                    "cells": [],
                    "rows": [],
                    "allRows": [],
                    "columns": [],
                    "headerRows": [],
                    "maxCol": 0,
                    "changedCount": 0,
                    "conflictCount": 0,
                    "structuralConflictCount": int(action.startswith("conflict_")),
                    "needsAttention": action.startswith("conflict_"),
                })
            return {
                "engineVersion": structural["version"],
                "sheets": sheets,
                "summary": {
                    "sheetCount": len(sheets),
                    "changedCellCount": sum(sheet["changedCount"] for sheet in sheets),
                    "cellConflictCount": sum(sheet["conflictCount"] for sheet in sheets),
                    "structuralConflictCount": sum(sheet.get("structuralConflictCount", 0) for sheet in sheets),
                    "sheetConflictCount": structural["summary"].get("sheetConflictCount", 0),
                },
            }

        sheets: list[dict[str, Any]] = []
        for name in _ordered_sheet_names(wb_base, wb_ours, wb_theirs):
            base_exists = wb_base is not None and name in wb_base.sheetnames
            ours_exists = name in wb_ours.sheetnames
            theirs_exists = name in wb_theirs.sheetnames
            ws_base = wb_base[name] if base_exists else None
            ws_ours = wb_ours[name] if ours_exists else None
            ws_theirs = wb_theirs[name] if theirs_exists else None
            ws_base_v = wb_base_v[name] if base_exists else None
            ws_ours_v = wb_ours_v[name] if ours_exists else None
            ws_theirs_v = wb_theirs_v[name] if theirs_exists else None

            if ours_exists and theirs_exists:
                sheet = _modified_sheet_diff(name, ws_base, ws_ours, ws_theirs,
                                             ws_base_v, ws_ours_v, ws_theirs_v)
                if sheet["changedCount"]:
                    sheets.append(sheet)
                continue
            if not base_exists:
                source = "ours" if ours_exists else "theirs"
                sheets.append({
                    "name": name,
                    "action": f"auto_add_{source}",
                    "cells": [],
                    "contextRows": {},
                    "maxCol": 0,
                    "changedCount": 0,
                    "conflictCount": 0,
                })
                continue
            if not ours_exists and not theirs_exists:
                sheets.append({
                    "name": name,
                    "action": "auto_delete",
                    "cells": [],
                    "contextRows": {},
                    "maxCol": 0,
                    "changedCount": 0,
                    "conflictCount": 0,
                })
                continue

            existing_ws = ws_ours or ws_theirs
            missing_side = "theirs" if ours_exists else "ours"
            if _sheet_equal(ws_base, existing_ws):
                action = "auto_delete"
                conflict_count = 0
            else:
                action = f"conflict_delete_{missing_side}"
                conflict_count = 1
            sheets.append({
                "name": name,
                "action": action,
                "cells": [],
                "contextRows": {},
                "maxCol": 0,
                "changedCount": 0,
                "conflictCount": conflict_count,
            })

        return {
            "sheets": sheets,
            "summary": {
                "sheetCount": len(sheets),
                "changedCellCount": sum(sheet["changedCount"] for sheet in sheets),
                "cellConflictCount": sum(
                    sum(cell["resolution"] == "conflict" for cell in sheet["cells"])
                    for sheet in sheets
                ),
                "sheetConflictCount": sum(sheet["action"].startswith("conflict_") for sheet in sheets),
            },
        }
    finally:
        if wb_base is not None:
            wb_base.close()
        wb_ours.close()
        wb_theirs.close()
        if wb_base_v is not None:
            wb_base_v.close()
        wb_ours_v.close()
        wb_theirs_v.close()


def _author_and_date(repo: Path, relative_path: str, ref: str) -> tuple[str, str]:
    verify = _run_git(repo, ["rev-parse", "--verify", "--quiet", ref])
    if verify.returncode != 0:
        return "", ""
    result = _run_git(repo, ["log", "-1", "--format=%an|%cI", ref, "--", relative_path])
    if result.returncode != 0 or not result.stdout.strip():
        return "", ""
    parts = result.stdout.strip().split("|", 1)
    if len(parts) != 2:
        return parts[0], ""
    author, iso_date = parts
    # ISO 8601 -> "2000-1-1 8:24"
    try:
        from datetime import datetime
        dt = datetime.fromisoformat(iso_date)
        date_str = f"{dt.year}-{dt.month}-{dt.day} {dt.hour}:{dt.minute:02d}"
    except (ValueError, TypeError):
        date_str = ""
    return author, date_str


def _author(repo: Path, relative_path: str, ref: str) -> str:
    author, _ = _author_and_date(repo, relative_path, ref)
    return author


def _theirs_author(repo: Path, relative_path: str) -> str:
    for ref in ("MERGE_HEAD", "CHERRY_PICK_HEAD", "REBASE_HEAD"):
        author = _author(repo, relative_path, ref)
        if author:
            return author
    return ""


def _ours_date(repo: Path, relative_path: str) -> str:
    _, date_str = _author_and_date(repo, relative_path, "HEAD")
    return date_str


def _theirs_date(repo: Path, relative_path: str) -> str:
    for ref in ("MERGE_HEAD", "CHERRY_PICK_HEAD", "REBASE_HEAD"):
        _, date_str = _author_and_date(repo, relative_path, ref)
        if date_str:
            return date_str
    return ""


def prepare_conflicts(
    repo: Path | str,
    *,
    relative_path: str | None = None,
    runtime_dir: Path | str | None = None,
) -> Path:
    repo_root = find_repo_root(repo)
    conflicts = detect_conflicts(repo_root)
    if relative_path:
        normalized = relative_path.replace("\\", "/")
        conflicts = [item for item in conflicts if item["path"] == normalized]
    if not conflicts:
        raise ConflictError("No unresolved .xlsx/.xlsm conflicts were found.")

    target_dir = _safe_runtime_dir(repo_root, runtime_dir)
    prepared_files = []
    for item in conflicts:
        path = item["path"]
        ours_blob = _git_blob(repo_root, 2, path)
        theirs_blob = _git_blob(repo_root, 3, path)
        if ours_blob is None or theirs_blob is None:
            raise ConflictError(
                f"{path} is a delete/modify conflict. This resolver currently requires both workbook versions."
            )
        base_blob = _git_blob(repo_root, 1, path)
        extension = Path(path).suffix.lower()
        macro = None
        if extension == ".xlsm":
            macro = _macro_payload_info(base_blob, ours_blob, theirs_blob)
        work_dir = target_dir / _slug(path)
        work_dir.mkdir(parents=True, exist_ok=True)
        ours_path = work_dir / f"ours{extension}"
        theirs_path = work_dir / f"theirs{extension}"
        base_path = work_dir / f"base{extension}"
        ours_path.write_bytes(ours_blob)
        theirs_path.write_bytes(theirs_blob)
        if base_blob is not None:
            base_path.write_bytes(base_blob)
        elif base_path.exists():
            base_path.unlink()
        workbook_diff = diff_workbooks(
            base_path if base_blob is not None else None,
            ours_path,
            theirs_path,
            analysis_dir=work_dir / "analysis",
        )
        prepared_files.append({
            "id": path,
            "path": path,
            "base": str(base_path) if base_blob is not None else None,
            "ours": str(ours_path),
            "theirs": str(theirs_path),
            "output": str((repo_root / Path(path)).resolve()),
            "oursAuthor": _author(repo_root, path, "HEAD"),
            "theirsAuthor": _theirs_author(repo_root, path),
            "oursDate": _ours_date(repo_root, path),
            "theirsDate": _theirs_date(repo_root, path),
            "macro": macro,
            "diff": workbook_diff,
        })

    manifest = {
        "version": 3,
        "repoRoot": str(repo_root),
        "files": prepared_files,
        "summary": {
            "fileCount": len(prepared_files),
            "cellConflictCount": sum(f["diff"]["summary"]["cellConflictCount"] for f in prepared_files),
            "sheetConflictCount": sum(f["diff"]["summary"]["sheetConflictCount"] for f in prepared_files),
            "structuralConflictCount": sum(f["diff"]["summary"].get("structuralConflictCount", 0) for f in prepared_files),
        },
    }
    manifest_path = target_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest_path


def load_manifest(path: Path | str) -> dict[str, Any]:
    manifest_path = Path(path).resolve()
    data = json.loads(manifest_path.read_text(encoding="utf-8"))
    if data.get("version") != 3 or not isinstance(data.get("files"), list):
        raise ConflictError(f"Unsupported manifest: {manifest_path}")
    return data


def _copy_cell(source, target) -> None:
    target.value = source.value
    if source.has_style:
        if source.parent is not None and source.parent.parent is target.parent.parent:
            # 同 workbook：样式表索引通用，浅拷贝 StyleArray 避免对象复制（大表 40 万+ 格）
            target._style = copy(source._style)
        else:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
    target.hyperlink = source.hyperlink
    target.comment = source.comment


def _copy_sheet(source_ws, target_wb, *, index: int | None = None):
    target_ws = target_wb.create_sheet(source_ws.title, index=index)
    for row in source_ws.iter_rows():
        for source_cell in row:
            if isinstance(source_cell, MergedCell):
                continue
            _copy_cell(source_cell, target_ws.cell(source_cell.row, source_cell.column))
    for key, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key] = copy(dimension)
        target_ws.column_dimensions[key].worksheet = target_ws
    for key, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key] = copy(dimension)
        target_ws.row_dimensions[key].worksheet = target_ws
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))
    target_ws.views = copy(source_ws.views)
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)
    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_state = source_ws.sheet_state
    return target_ws


def _replace_sheet(target_wb, source_ws, order: list[str]) -> None:
    if source_ws.title in target_wb.sheetnames:
        index = target_wb.sheetnames.index(source_ws.title)
        target_wb.remove(target_wb[source_ws.title])
    else:
        desired = order.index(source_ws.title) if source_ws.title in order else len(target_wb.sheetnames)
        index = min(desired, len(target_wb.sheetnames))
    _copy_sheet(source_ws, target_wb, index=index)


def _file_decisions(decisions: dict[str, Any], file_id: str) -> dict[str, Any]:
    return ((decisions.get("files") or {}).get(file_id) or {})


def _cell_choice(cell_choices: dict[str, Any], cell: dict[str, Any]) -> Any:
    return cell_choices.get(cell.get("id"), cell_choices.get(cell.get("coordinate")))


def _validate_decisions(manifest: dict[str, Any], decisions: dict[str, Any]) -> None:
    missing: list[str] = []
    for file_entry in manifest["files"]:
        file_choice = _file_decisions(decisions, file_entry["id"])
        macro = file_entry.get("macro") or {}
        macro_choice = file_choice.get("macro", macro.get("default"))
        if macro.get("status") == "conflict" and macro_choice not in ("ours", "theirs"):
            missing.append(f"{file_entry['path']} / <VBA/ActiveX>")
        if macro_choice == "theirs" and not macro.get("canChooseTheirs", True):
            raise ConflictError(
                f"{file_entry['path']} has incompatible LOCAL and REMOTE VBA/ActiveX package structures; "
                "only the LOCAL macro payload can be preserved safely."
            )
        sheet_choices = file_choice.get("sheets") or {}
        for sheet in file_entry["diff"]["sheets"]:
            choice = sheet_choices.get(sheet["name"]) or {}
            if sheet["action"].startswith("conflict_") and choice.get("sheet") not in ("ours", "theirs"):
                missing.append(f"{file_entry['path']} / {sheet['name']} / <sheet>")
            if sheet.get("renameStatus") == "rename_conflict" and choice.get("rename") not in ("ours", "theirs"):
                missing.append(f"{file_entry['path']} / {sheet['name']} / <rename>")
            if (sheet.get("mergedCells") or {}).get("status") == "conflict" and choice.get("mergedCells") not in ("ours", "theirs"):
                missing.append(f"{file_entry['path']} / {sheet['name']} / <merged-cells>")
            row_choices = choice.get("rows") or {}
            for row in sheet.get("allRows") or sheet.get("rows") or []:
                if row.get("state") == "delete_modify" and row_choices.get(row["id"]) not in ("keep", "delete"):
                    missing.append(f"{file_entry['path']} / {sheet['name']} / {row['id']}")
            column_choices = choice.get("columns") or {}
            for column in sheet.get("columns") or []:
                if column.get("state") == "delete_modify" and column_choices.get(column["id"]) not in ("keep", "delete"):
                    missing.append(f"{file_entry['path']} / {sheet['name']} / {column['id']}")
            cell_choices = choice.get("cells") or {}
            for cell in sheet["cells"]:
                if cell["resolution"] == "conflict" and _cell_choice(cell_choices, cell) not in ("ours", "theirs"):
                    missing.append(f"{file_entry['path']} / {sheet['name']} / {cell['coordinate']}")
    if missing:
        preview = "\n".join(f"- {item}" for item in missing[:20])
        suffix = f"\n... and {len(missing) - 20} more" if len(missing) > 20 else ""
        raise ConflictError(f"Unresolved choices remain:\n{preview}{suffix}")


def _copy_sheet_properties(source_ws, target_ws) -> None:
    # freeze_panes only exposes topLeftCell. Reassigning it turns a scrolled
    # viewport such as topLeftCell=D857/ySplit=4 into a 856-row freeze.
    target_ws.views = copy(source_ws.views)
    target_ws.sheet_format = copy(source_ws.sheet_format)
    target_ws.sheet_properties = copy(source_ws.sheet_properties)
    target_ws.page_margins = copy(source_ws.page_margins)
    target_ws.page_setup = copy(source_ws.page_setup)
    target_ws.print_options = copy(source_ws.print_options)
    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.sheet_state = source_ws.sheet_state


def _restore_merged_cells(target_ws, sheet, selection, active_rows, active_columns, header_count) -> None:
    model = sheet.get("mergedCells") or {}
    if model.get("status") == "conflict":
        regions = model.get(selection.get("mergedCells")) or []
    else:
        regions = model.get("resolved") or []
    row_positions = {f"header:{row}": row for row in range(1, header_count + 1)}
    row_positions.update({row["id"]: header_count + index for index, row in enumerate(active_rows, 1)})
    column_positions = {column["id"]: index for index, column in enumerate(active_columns, 1)}
    occupied: set[tuple[int, int]] = set()
    for region in regions:
        output_rows = sorted({row_positions[row] for row in region["rows"] if row in row_positions})
        output_columns = sorted({column_positions[column] for column in region["columns"] if column in column_positions})
        if not output_rows or not output_columns or len(output_rows) * len(output_columns) < 2:
            continue
        if output_rows != list(range(output_rows[0], output_rows[-1] + 1)):
            raise ConflictError(f"Merged rows are no longer contiguous in {sheet['name']}")
        if output_columns != list(range(output_columns[0], output_columns[-1] + 1)):
            raise ConflictError(f"Merged columns are no longer contiguous in {sheet['name']}")
        cells = {(row, column) for row in output_rows for column in output_columns}
        if occupied & cells:
            raise ConflictError(f"Merged regions overlap after alignment in {sheet['name']}")
        occupied.update(cells)
        target_ws.merge_cells(
            start_row=output_rows[0], end_row=output_rows[-1],
            start_column=output_columns[0], end_column=output_columns[-1],
        )


def _axis_kept(axis: dict[str, Any], choices: dict[str, Any]) -> bool:
    if axis.get("state") == "deleted":
        return False
    if axis.get("state") == "delete_modify":
        return choices.get(axis["id"], axis.get("default", "keep")) == "keep"
    return True


def _source_cell(workbooks, names, row, column, side: str):
    workbook = workbooks.get(side)
    sheet_name = names.get(side)
    row_number = row.get(f"{side}Row")
    column_number = column.get(f"{side}Col")
    if workbook is None or sheet_name is None or row_number is None or column_number is None:
        return None
    return workbook[sheet_name].cell(row_number, column_number)


def _selected_source_cell(workbooks, names, row, column, cell, choice: str | None):
    preferred = []
    if choice in ("ours", "theirs"):
        preferred.append(choice)
    elif cell.get("resolution") == "theirs":
        preferred.append("theirs")
    elif cell.get("resolution") == "ours":
        preferred.append("ours")
    elif cell.get("resolution") == "same":
        preferred.extend(("ours", "theirs"))
    else:
        preferred.append("theirs")
    preferred.extend(side for side in ("ours", "theirs", "base") if side not in preferred)
    for side in preferred:
        source = _source_cell(workbooks, names, row, column, side)
        if source is not None:
            return source
    return None


def _selected_sheet_name(sheet: dict[str, Any], selection: dict[str, Any]) -> str:
    names = sheet.get("names") or {}
    if sheet.get("renameStatus") == "rename_conflict":
        return names[selection["rename"]]
    return names.get("ours") or names.get("theirs") or names.get("base") or sheet["name"]


_FORMULA_CACHE_NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
_FORMULA_CACHE_RNS = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'


def _inject_formula_cache(temp_path: Path, cache: dict[str, dict[str, Any]]) -> None:
    """后处理 xlsx XML，为公式 cell 注入缓存值，使 data_only 读取得到正确值。"""
    if not cache:
        return
    import xml.etree.ElementTree as ET
    ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
    with zipfile.ZipFile(temp_path) as zin:
        infos = zin.infolist()
        parts = {info.filename: zin.read(info.filename) for info in infos}
    wb_tree = ET.fromstring(parts['xl/workbook.xml'])
    rels_tree = ET.fromstring(parts['xl/_rels/workbook.xml.rels'])
    rels = {rel.get('Id'): rel.get('Target') for rel in rels_tree}
    sheet_to_file = {}
    for sheet in wb_tree.iter(f'{_FORMULA_CACHE_NS}sheet'):
        name = sheet.get('name')
        rel_id = sheet.get(f'{_FORMULA_CACHE_RNS}id')
        target = rels.get(rel_id, '')
        if target:
            sheet_to_file[name] = target.lstrip('/') if target.startswith('/') else 'xl/' + target
    for sheet_name, coord_cache in cache.items():
        xml_path = sheet_to_file.get(sheet_name)
        if not xml_path or xml_path not in parts:
            continue
        tree = ET.fromstring(parts[xml_path])
        modified = False
        for c in tree.iter(f'{_FORMULA_CACHE_NS}c'):
            coord = c.get('r')
            if coord not in coord_cache:
                continue
            if c.find(f'{_FORMULA_CACHE_NS}f') is None:
                continue
            cached = coord_cache[coord]
            if cached is None:
                continue
            v_elem = c.find(f'{_FORMULA_CACHE_NS}v')
            if v_elem is None:
                v_elem = ET.SubElement(c, f'{_FORMULA_CACHE_NS}v')
            if isinstance(cached, bool):
                c.set('t', 'b')
                v_elem.text = '1' if cached else '0'
            elif isinstance(cached, str):
                c.set('t', 'str')
                v_elem.text = cached
            else:
                v_elem.text = str(cached)
            modified = True
        if modified:
            xml_bytes = ET.tostring(tree, encoding='UTF-8')
            if xml_bytes.startswith(b"<?xml"):
                xml_bytes = xml_bytes.replace(b"encoding='UTF-8'", b'encoding="UTF-8"')
                if b'standalone' not in xml_bytes:
                    xml_bytes = xml_bytes.replace(b'?>', b' standalone="yes"?>', 1)
            else:
                xml_bytes = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + xml_bytes
            parts[xml_path] = xml_bytes
    with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for info in infos:
            zout.writestr(info, parts[info.filename])


def _rebuild_structural_sheet(wb_out, wb_base, wb_ours, wb_theirs, sheet, selection,
                              wb_base_v=None, wb_ours_v=None, wb_theirs_v=None) -> dict[str, Any]:
    names = sheet.get("names") or {"base": sheet["name"], "ours": sheet["name"], "theirs": sheet["name"]}
    source_ws = None
    for workbook, side in ((wb_ours, "ours"), (wb_theirs, "theirs"), (wb_base, "base")):
        if workbook is not None and names.get(side) in workbook.sheetnames:
            source_ws = workbook[names[side]]
            break
    if source_ws is None:
        raise ConflictError(f"No source sheet is available for {sheet['name']}")

    existing_name = names.get("ours")
    index = wb_out.sheetnames.index(existing_name) if existing_name in wb_out.sheetnames else len(wb_out.sheetnames)
    temp_title = f"__merge_{len(wb_out.sheetnames)}__"
    while temp_title in wb_out.sheetnames:
        temp_title += "_"
    target_ws = wb_out.create_sheet(temp_title, index=index)
    _copy_sheet_properties(source_ws, target_ws)

    row_choices = selection.get("rows") or {}
    column_choices = selection.get("columns") or {}
    active_rows = [row for row in sheet.get("allRows") or [] if _axis_kept(row, row_choices)]
    active_columns = [column for column in sheet.get("columns") or [] if _axis_kept(column, column_choices)]
    workbooks = {"base": wb_base, "ours": wb_ours, "theirs": wb_theirs}
    workbooks_v = {"base": wb_base_v, "ours": wb_ours_v, "theirs": wb_theirs_v}
    cell_map = {cell["id"]: cell for cell in sheet.get("allCells") or sheet.get("cells") or []}
    cell_choices = selection.get("cells") or {}
    formula_cache: dict[str, Any] = {}

    header_count = max(
        [max(_header_rows(workbook[names[side]]) or [0])
         for workbook, side in ((wb_base, "base"), (wb_ours, "ours"), (wb_theirs, "theirs"))
         if workbook is not None and names.get(side) in workbook.sheetnames]
        or [0]
    )
    for output_col, column in enumerate(active_columns, 1):
        dimension_source = None
        for workbook, side in ((wb_ours, "ours"), (wb_theirs, "theirs"), (wb_base, "base")):
            source_col = column.get(f"{side}Col")
            if workbook is not None and names.get(side) in workbook.sheetnames and source_col is not None:
                dimension_source = workbook[names[side]].column_dimensions[get_column_letter(source_col)]
                break
        if dimension_source is not None:
            target_ws.column_dimensions[get_column_letter(output_col)] = copy(dimension_source)
            target_ws.column_dimensions[get_column_letter(output_col)].worksheet = target_ws
    # 表头行高：ours 优先，回退 theirs/base
    for header_row in range(1, header_count + 1):
        for workbook, side in ((wb_ours, "ours"), (wb_theirs, "theirs"), (wb_base, "base")):
            if workbook is not None and names.get(side) in workbook.sheetnames:
                dim = workbook[names[side]].row_dimensions[header_row]
                if dim.height is not None:
                    target_ws.row_dimensions[header_row] = copy(dim)
                    target_ws.row_dimensions[header_row].worksheet = target_ws
                break
    for header_row in range(1, header_count + 1):
        for output_col, column in enumerate(active_columns, 1):
            source = None
            for side in ("theirs", "ours", "base"):
                workbook = workbooks[side]
                source_col = column.get(f"{side}Col")
                if workbook is not None and names.get(side) in workbook.sheetnames and source_col is not None:
                    source = workbook[names[side]].cell(header_row, source_col)
                    if source.value is not None:
                        break
            if source is not None:
                _copy_cell(source, target_ws.cell(header_row, output_col))

    for output_offset, row in enumerate(active_rows, 1):
        output_row = header_count + output_offset
        dimension_source = None
        for workbook, side in ((wb_ours, "ours"), (wb_theirs, "theirs"), (wb_base, "base")):
            source_row = row.get(f"{side}Row")
            if workbook is not None and names.get(side) in workbook.sheetnames and source_row is not None:
                dimension_source = workbook[names[side]].row_dimensions[source_row]
                break
        if dimension_source is not None:
            target_ws.row_dimensions[output_row] = copy(dimension_source)
            target_ws.row_dimensions[output_row].worksheet = target_ws
        for output_col, column in enumerate(active_columns, 1):
            cell = cell_map.get(f'{row["id"]}|{column["id"]}') or {
                "resolution": "same", "id": f'{row["id"]}|{column["id"]}', "coordinate": ""
            }
            choice = _cell_choice(cell_choices, cell)
            source = _selected_source_cell(workbooks, names, row, column, cell, choice)
            if source is not None:
                # 空且无格式的 cell 跳过写入：目标默认状态即空，输出等价
                if (
                    source.value is None
                    and not source.has_style
                    and source.hyperlink is None
                    and source.comment is None
                ):
                    continue
                _copy_cell(source, target_ws.cell(output_row, output_col))
                # 公式 cell：记录输出坐标 -> 缓存值，保存后注入 XML
                if source.data_type == 'f':
                    cached_source = _selected_source_cell(workbooks_v, names, row, column, cell, choice)
                    cached_val = cached_source.value if cached_source is not None else None
                    if cached_val is not None:
                        formula_cache[f'{get_column_letter(output_col)}{output_row}'] = cached_val

    _restore_merged_cells(target_ws, sheet, selection, active_rows, active_columns, header_count)

    if existing_name in wb_out.sheetnames:
        wb_out.remove(wb_out[existing_name])
    final_name = _selected_sheet_name(sheet, selection)
    if final_name in wb_out.sheetnames and wb_out[final_name] is not target_ws:
        raise ConflictError(f"Sheet name collision after merge: {final_name}")
    target_ws.title = final_name
    return {final_name: formula_cache}


def _build_merged_file(file_entry: dict[str, Any], decisions: dict[str, Any]) -> Path:
    output = Path(file_entry["output"]).resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    extension = output.suffix.lower()
    keep_vba = extension == ".xlsm"
    file_choice = _file_decisions(decisions, file_entry["id"])
    macro = file_entry.get("macro") or {}
    macro_choice = file_choice.get("macro", macro.get("default", "ours"))
    fd, temp_name = tempfile.mkstemp(prefix=f".{output.stem}-", suffix=extension, dir=output.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    wb_out = load_workbook(file_entry["ours"], data_only=False, keep_links=True, keep_vba=keep_vba)
    wb_base = load_workbook(file_entry["base"], data_only=False, keep_links=True) if file_entry.get("base") else None
    wb_ours = wb_out  # 只读复用，使 _copy_cell 同 workbook 样式索引快路径生效
    wb_theirs = load_workbook(file_entry["theirs"], data_only=False, keep_links=True)
    wb_base_v = load_workbook(file_entry["base"], data_only=True, keep_links=True) if file_entry.get("base") else None
    wb_ours_v = load_workbook(file_entry["ours"], data_only=True, keep_links=True)
    wb_theirs_v = load_workbook(file_entry["theirs"], data_only=True, keep_links=True)
    sheet_choices = file_choice.get("sheets") or {}
    formula_cache: dict[str, dict[str, Any]] = {}
    try:
        theirs_order = wb_theirs.sheetnames
        for sheet in file_entry["diff"]["sheets"]:
            name = sheet["name"]
            action = sheet["action"]
            selection = sheet_choices.get(name) or {}
            if action == "auto_delete":
                if name in wb_out.sheetnames:
                    wb_out.remove(wb_out[name])
                continue
            if action == "auto_add_theirs":
                theirs_name = (sheet.get("names") or {}).get("theirs") or name
                ours_name = (sheet.get("names") or {}).get("ours")
                if ours_name in wb_out.sheetnames:
                    wb_out.remove(wb_out[ours_name])
                if theirs_name not in wb_out.sheetnames:
                    _copy_sheet(wb_theirs[theirs_name], wb_out, index=min(theirs_order.index(theirs_name), len(wb_out.sheetnames)))
                continue
            if action.startswith("conflict_"):
                side = selection["sheet"]
                if side == "ours":
                    continue
                if name in wb_theirs.sheetnames:
                    _replace_sheet(wb_out, wb_theirs[name], theirs_order)
                elif name in wb_out.sheetnames:
                    wb_out.remove(wb_out[name])
                continue
            if action != "merge":
                continue
            if not sheet.get("rebuildRequired", True):
                existing_name = (sheet.get("names") or {}).get("ours") or name
                final_name = _selected_sheet_name(sheet, selection)
                if existing_name in wb_out.sheetnames and final_name != existing_name:
                    if final_name in wb_out.sheetnames:
                        raise ConflictError(f"Sheet name collision after merge: {final_name}")
                    wb_out[existing_name].title = final_name
                continue
            if sheet.get("allRows") is not None and sheet.get("columns") is not None:
                cache = _rebuild_structural_sheet(
                    wb_out, wb_base, wb_ours, wb_theirs, sheet, selection,
                    wb_base_v, wb_ours_v, wb_theirs_v,
                )
                formula_cache.update(cache)
                continue
            ws_out = wb_out[name]
            ws_theirs = wb_theirs[name]
            ws_theirs_v = wb_theirs_v[name] if name in wb_theirs_v.sheetnames else None
            cell_choices = selection.get("cells") or {}
            legacy_cache: dict[str, Any] = {}
            for cell in sheet["cells"]:
                side = cell_choices.get(cell["coordinate"], cell["resolution"])
                if side == "conflict":
                    side = cell_choices[cell["coordinate"]]
                if side == "theirs":
                    src = ws_theirs[cell["coordinate"]]
                    _copy_cell(src, ws_out[cell["coordinate"]])
                    if src.data_type == 'f' and ws_theirs_v is not None:
                        cached = ws_theirs_v[cell["coordinate"]].value
                        if cached is not None:
                            legacy_cache[cell["coordinate"]] = cached
            if legacy_cache:
                formula_cache[name] = legacy_cache
        if not wb_out.sheetnames:
            raise ConflictError(f"Cannot save a workbook with no sheets: {file_entry['path']}")
        wb_out.save(temp_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise
    finally:
        _close_workbook(wb_out)
        _close_workbook(wb_base)
        if wb_ours is not wb_out:
            _close_workbook(wb_ours)
        _close_workbook(wb_theirs)
        _close_workbook(wb_base_v)
        _close_workbook(wb_ours_v)
        _close_workbook(wb_theirs_v)

    _inject_formula_cache(temp_path, formula_cache)
    check_wb = load_workbook(temp_path, read_only=True, data_only=False)
    check_wb.close()
    if keep_vba:
        macro_source = file_entry["theirs"] if macro_choice == "theirs" else file_entry["ours"]
        if macro_choice == "theirs":
            if not macro.get("canChooseTheirs", True):
                temp_path.unlink(missing_ok=True)
                raise ConflictError(
                    f"Cannot safely preserve the REMOTE macro payload for {file_entry['path']} because its package structure differs."
                )
            _replace_macro_payload(temp_path, macro_source)
        expected_hash, expected_entries = _macro_payload_digest(macro_source)
        actual_hash, actual_entries = _macro_payload_digest(temp_path)
        if actual_hash != expected_hash or actual_entries != expected_entries:
            temp_path.unlink(missing_ok=True)
            raise ConflictError(f"Macro payload changed while saving {file_entry['path']}; output was not written.")
    return temp_path


def _git_step(repo: Path, args: list[str], label: str) -> dict[str, Any]:
    result = _run_git(repo, args)
    output = (result.stdout + result.stderr).strip()
    return {"step": label, "success": result.returncode == 0, "output": output}


def _build_commit_message(manifest: dict[str, Any], decisions: dict[str, Any]) -> str:
    """生成结构化提交信息：每个已选择的冲突 Cell 一行。

    格式：resolve: [表:path]-[sheet:name]-[单元格:A1]-[选择:远端/本地]  [远端值:xxx  本地值:xxx]
    无冲突 Cell 时回退默认信息。
    """
    lines: list[str] = []
    for file_entry in manifest["files"]:
        file_choice = _file_decisions(decisions, file_entry["id"])
        sheet_choices = file_choice.get("sheets") or {}
        for sheet in file_entry["diff"]["sheets"]:
            choice = sheet_choices.get(sheet["name"]) or {}
            cell_choices = choice.get("cells") or {}
            for cell in sheet["cells"]:
                if cell.get("resolution") != "conflict":
                    continue
                cell_choice = _cell_choice(cell_choices, cell)
                if cell_choice not in ("ours", "theirs"):
                    continue
                label = cell.get("coordinate") or cell.get("id") or ""
                side = "远端" if cell_choice == "theirs" else "本地"
                theirs_value = _message_value(cell.get("theirs"))
                ours_value = _message_value(cell.get("ours"))
                lines.append(
                    f"resolve: [表:{file_entry['path']}]-[sheet:{sheet['name']}]-[单元格:{label}]-[选择:{side}]  "
                    f"[远端值:{theirs_value}  本地值:{ours_value}]"
                )
    if not lines:
        return "resolve: merge xlsx conflicts"
    message = "\n".join(lines)
    # 防止超长消息撑爆 git commit 命令行（Windows 单条命令约 32K 字符）
    if len(message) > MAX_COMMIT_MESSAGE_CHARS:
        message = message[:MAX_COMMIT_MESSAGE_CHARS].rstrip() + "\n... (truncated)"
    return message


MAX_COMMIT_MESSAGE_CHARS = 20000


def _message_value(value) -> str:
    """Cell 显示值转提交信息字符串：None 转空，浮点去多余小数。"""
    if value is None:
        return ""
    if isinstance(value, float):
        return str(int(value)) if value == int(value) else str(value)
    return str(value)


def apply_manifest(
    manifest_path: Path | str,
    decisions: dict[str, Any],
    *,
    commit: bool = True,
    push: bool = False,
    message: str | None = None,
) -> dict[str, Any]:
    manifest = load_manifest(manifest_path)
    _validate_decisions(manifest, decisions)
    repo = find_repo_root(manifest["repoRoot"])
    prepared: list[tuple[Path, Path]] = []
    try:
        for file_entry in manifest["files"]:
            output = Path(file_entry["output"]).resolve()
            if os.path.commonpath([str(repo), str(output)]) != str(repo):
                raise ConflictError(f"Output escapes meta repository: {output}")
            prepared.append((output, _build_merged_file(file_entry, decisions)))
        for output, temp_path in prepared:
            os.replace(temp_path, output)

        relative_paths = [str(output.relative_to(repo)).replace("\\", "/") for output, _ in prepared]
        steps = [_git_step(repo, ["add", "--", *relative_paths], "git add")]
        success = steps[-1]["success"]
        if success and commit:
            commit_message = message or _build_commit_message(manifest, decisions)
            steps.append(_git_step(repo, ["commit", "-m", commit_message], "git commit"))
            success = steps[-1]["success"]
        if success and push:
            steps.append(_git_step(repo, ["push"], "git push"))
            success = steps[-1]["success"]
        return {
            "success": success,
            "merged": [str(output) for output, _ in prepared],
            "steps": steps,
        }
    finally:
        for _, temp_path in prepared:
            temp_path.unlink(missing_ok=True)
