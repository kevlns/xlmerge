import json
import os
import subprocess
import sys
import tempfile
import time
import unittest
import urllib.request
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.views import Pane


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
RESOLVER_DIR = PACKAGE_ROOT / "xlsx_resolver"
sys.path.insert(0, str(RESOLVER_DIR))

from xlsx_conflict import (
    ConflictError,
    _build_merged_file,
    _close_workbook,
    apply_manifest,
    detect_conflicts,
    diff_workbooks,
    load_manifest,
    prepare_conflicts,
)


def git(repo, *args, check=True):
    result = subprocess.run(
        ["git", "-C", str(repo), *args],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if check and result.returncode:
        raise AssertionError(result.stdout + result.stderr)
    return result


def write_book(path, version):
    wb = Workbook()
    wb.remove(wb.active)

    alpha = wb.create_sheet("Alpha")
    alpha.append(["##var", "value"])
    alpha.append(["id", {"base": "base-alpha", "ours": "ours-alpha", "theirs": "theirs-alpha"}[version]])

    beta = wb.create_sheet("Beta")
    beta["A1"] = {"base": "base-beta", "ours": "ours-beta", "theirs": "theirs-beta"}[version]

    auto = wb.create_sheet("Auto")
    auto["A1"] = "ours-auto" if version == "ours" else "base-auto"
    auto["B1"] = "=1+2" if version == "theirs" else "base-formula"
    if version == "theirs":
        auto["B1"].fill = PatternFill("solid", fgColor="00FF00")

    if version != "ours":
        delete_modify = wb.create_sheet("DeleteModify")
        delete_modify["A1"] = "changed-theirs" if version == "theirs" else "base-delete"

    if version == "ours":
        wb.create_sheet("OursOnly")["A1"] = "ours-only"
    if version == "theirs":
        wb.create_sheet("TheirsOnly")["A1"] = "theirs-only"
    wb.save(path)
    wb.close()


def write_grid(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def write_named_grid(path, sheet_name, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def write_luban_anonymous_book(path, value, *, note="行军队列", pane_top_left=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["##var", "id", None, "value"])
    ws.append(["##type", "int", None, "string"])
    ws.append(["##group", None, None, "c"])
    ws.append(["##", "ID", "策划注释", "值"])
    ws.append([None, 1, note, value])
    ws.column_dimensions["C"].width = 19.25
    ws["C5"].fill = PatternFill("solid", fgColor="FFF2CC")
    if pane_top_left:
        ws.sheet_view.pane = Pane(
            xSplit=3, ySplit=4, topLeftCell=pane_top_left,
            activePane="bottomRight", state="frozen",
        )
    wb.save(path)
    wb.close()


def write_structured_header_book(path, value, *, merged_range="I1:M1", include_aux=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append([
        "##var", "id", "value", None, None, None, None, None,
        "*items", None, None, None, None, None, "tail",
    ])
    ws.append([
        "##type", "int", "string", None, None, None, None, None,
        "name", "alias", "value", "comment", "tags", "extra", "string",
    ])
    ws.append([
        "##", "ID", "值", None, None, None, None, None,
        "枚举名", "别名", "枚举值", "注释", "标签", "扩展", "尾列",
    ])
    ws.append([
        None, 1, value, None, None, None, None, None,
        "entry", "条目", 1, "comment", "tag", "extra", "tail",
    ])
    ws.merge_cells(merged_range)
    if include_aux:
        aux = wb.create_sheet("Aux")
        aux["A1"] = "left"
        aux["C1"] = "right"
        aux["E1"] = "group"
        aux.merge_cells("E1:G1")
        aux["P10"].fill = PatternFill("solid", fgColor="FFF2CC")
    wb.save(path)
    wb.close()


def write_macro_book(path, value, macro_payload):
    source = Path(path).with_name(Path(path).name + ".source.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = value
    wb.save(source)
    wb.close()
    with zipfile.ZipFile(source) as archive:
        parts = {name: archive.read(name) for name in archive.namelist()}
    source.unlink()
    parts["[Content_Types].xml"] = parts["[Content_Types].xml"].replace(
        b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    ).replace(
        b"</Types>",
        b'<Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/></Types>',
    )
    parts["xl/_rels/workbook.xml.rels"] = parts["xl/_rels/workbook.xml.rels"].replace(
        b"</Relationships>",
        b'<Relationship Id="rIdVba" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/></Relationships>',
    )
    parts["xl/vbaProject.bin"] = bytes(macro_payload)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in parts.items():
            archive.writestr(name, content)


def read_macro_payload(path):
    with zipfile.ZipFile(path) as archive:
        return archive.read("xl/vbaProject.bin")


def _pid_exists(pid):
    """Cross-platform liveness probe (tasklist is Windows-only; CI runs on ubuntu)."""
    if sys.platform == "win32":
        result = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}", "/FO", "CSV", "/NH"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        return f'"{pid}"' in result.stdout
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        return True  # exists but owned by another user
    return True


def wait_for_process_exit(pid, timeout=5):
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if not _pid_exists(pid):
            return
        time.sleep(0.1)
    raise AssertionError(f"Resolver process {pid} did not exit after shutdown")


class XlsxConflictIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.repo = Path(self.temp.name) / "meta"
        self.repo.mkdir()
        git(self.repo, "init")
        git(self.repo, "config", "user.name", "Resolver Test")
        git(self.repo, "config", "user.email", "resolver@example.invalid")
        self.path = self.repo / "new_meta" / "MultiSheet.xlsx"
        self.path.parent.mkdir()
        write_book(self.path, "base")
        git(self.repo, "add", ".")
        git(self.repo, "commit", "-m", "base")
        git(self.repo, "checkout", "-b", "theirs")
        write_book(self.path, "theirs")
        git(self.repo, "add", ".")
        git(self.repo, "commit", "-m", "theirs")
        git(self.repo, "checkout", "-")
        write_book(self.path, "ours")
        git(self.repo, "add", ".")
        git(self.repo, "commit", "-m", "ours")
        merge = git(self.repo, "merge", "theirs", check=False)
        self.assertNotEqual(merge.returncode, 0)

    def tearDown(self):
        self.temp.cleanup()

    def test_multi_sheet_diff_requires_file_sheet_cell_identity(self):
        conflicts = detect_conflicts(self.repo)
        self.assertEqual(conflicts, [{"path": "new_meta/MultiSheet.xlsx", "stages": [1, 2, 3]}])
        manifest_path = prepare_conflicts(self.repo, runtime_dir=Path(self.temp.name) / "runtime")
        manifest = load_manifest(manifest_path)
        file_entry = manifest["files"][0]
        sheets = {sheet["name"]: sheet for sheet in file_entry["diff"]["sheets"]}

        self.assertEqual(file_entry["diff"]["summary"]["cellConflictCount"], 2)
        self.assertEqual(sheets["Alpha"]["cells"][0]["coordinate"], "B2")
        self.assertEqual(sheets["Alpha"]["cells"][0]["resolution"], "conflict")
        self.assertEqual(sheets["Beta"]["cells"][0]["coordinate"], "A1")
        self.assertEqual(sheets["Beta"]["cells"][0]["resolution"], "conflict")
        self.assertEqual(sheets["DeleteModify"]["action"], "conflict_delete_ours")
        self.assertEqual(sheets["TheirsOnly"]["action"], "auto_add_theirs")

    def test_selected_cells_auto_changes_and_sheets_are_written_and_committed(self):
        manifest_path = prepare_conflicts(self.repo, runtime_dir=Path(self.temp.name) / "runtime")
        decisions = {
            "files": {
                "new_meta/MultiSheet.xlsx": {
                    "sheets": {
                        "Alpha": {"cells": {"B2": "ours"}},
                        "Beta": {"cells": {"A1": "theirs"}},
                        "DeleteModify": {"sheet": "theirs", "cells": {}},
                    }
                }
            }
        }
        result = apply_manifest(manifest_path, decisions, commit=True, message="resolve multi-sheet workbook")
        self.assertTrue(result["success"], result)
        self.assertEqual(git(self.repo, "diff", "--name-only", "--diff-filter=U").stdout.strip(), "")
        self.assertEqual(git(self.repo, "log", "-1", "--format=%s").stdout.strip(), "resolve multi-sheet workbook")

        wb = load_workbook(self.path, data_only=False)
        try:
            self.assertEqual(wb["Alpha"]["B2"].value, "ours-alpha")
            self.assertEqual(wb["Beta"]["A1"].value, "theirs-beta")
            self.assertEqual(wb["Auto"]["A1"].value, "ours-auto")
            self.assertEqual(wb["Auto"]["B1"].value, "=1+2")
            self.assertEqual(wb["Auto"]["B1"].fill.fgColor.rgb, "0000FF00")
            self.assertEqual(wb["DeleteModify"]["A1"].value, "changed-theirs")
            self.assertEqual(wb["OursOnly"]["A1"].value, "ours-only")
            self.assertEqual(wb["TheirsOnly"]["A1"].value, "theirs-only")
        finally:
            wb.close()

    def test_missing_sheet_qualified_choice_does_not_touch_worktree(self):
        manifest_path = prepare_conflicts(self.repo, runtime_dir=Path(self.temp.name) / "runtime")
        before = self.path.read_bytes()
        with self.assertRaises(ConflictError):
            apply_manifest(manifest_path, {"files": {}}, commit=False)
        self.assertEqual(self.path.read_bytes(), before)
        self.assertTrue(git(self.repo, "ls-files", "-u").stdout.strip())

    def test_local_http_api_exposes_all_sheets_and_applies_sheet_scoped_choices(self):
        runtime = Path(self.temp.name) / "runtime"
        entry = RESOLVER_DIR / "resolve_xlsx_conflict.py"
        process = subprocess.Popen(
            [
                sys.executable,
                str(entry),
                "--repo",
                str(self.repo),
                "--runtime-dir",
                str(runtime),
                "resolve",
                "--no-browser",
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
        )
        try:
            line = process.stdout.readline().strip()
            self.assertTrue(line.startswith("MERGE_SERVER_URL="), line)
            url = line.split("=", 1)[1]
            with urllib.request.urlopen(f"{url}/api/diff", timeout=5) as response:
                payload = json.load(response)
            sheet_names = [sheet["name"] for sheet in payload["files"][0]["diff"]["sheets"]]
            self.assertIn("Alpha", sheet_names)
            self.assertIn("Beta", sheet_names)

            body = json.dumps({
                "decisions": {
                    "files": {
                        "new_meta/MultiSheet.xlsx": {
                            "sheets": {
                                "Alpha": {"cells": {"B2": "theirs"}},
                                "Beta": {"cells": {"A1": "ours"}},
                                "Auto": {"cells": {"A1": "theirs", "B1": "ours"}},
                                "DeleteModify": {"sheet": "ours", "cells": {}},
                            }
                        }
                    }
                },
                "commit": False,
                "push": False,
            }).encode("utf-8")

            # 只读预览：返回结构化 commit 信息，不写盘、不提交
            preview_body = json.dumps({
                "decisions": {
                    "files": {
                        "new_meta/MultiSheet.xlsx": {
                            "sheets": {
                                "Alpha": {"cells": {"B2": "theirs"}},
                                "Beta": {"cells": {"A1": "ours"}},
                            }
                        }
                    }
                }
            }).encode("utf-8")
            preview_request = urllib.request.Request(
                f"{url}/api/commit-preview",
                data=preview_body,
                method="POST",
                headers={"Content-Type": "application/json"},
            )
            with urllib.request.urlopen(preview_request, timeout=5) as response:
                preview = json.load(response)
            self.assertIn("resolve: [表:new_meta/MultiSheet.xlsx]-[sheet:Alpha]-[单元格:", preview["message"])
            self.assertIn("[选择:远端]", preview["message"])  # Alpha B2 → theirs
            self.assertIn("[选择:本地]", preview["message"])  # Beta A1 → ours
            self.assertIn("[远端值:theirs-beta  本地值:ours-beta]", preview["message"])

            request = urllib.request.Request(
                f"{url}/api/resolve",
                data=body,
                method="POST",
                headers={"Content-Type": "application/json"},
            )
            with urllib.request.urlopen(request, timeout=10) as response:
                result = json.load(response)
            self.assertTrue(result["success"], result)
            process.wait(timeout=10)

            wb = load_workbook(self.path, data_only=False)
            try:
                self.assertEqual(wb["Alpha"]["B2"].value, "theirs-alpha")
                self.assertEqual(wb["Beta"]["A1"].value, "ours-beta")
                self.assertEqual(wb["Auto"]["A1"].value, "base-auto")
                self.assertEqual(wb["Auto"]["B1"].value, "base-formula")
                self.assertNotIn("DeleteModify", wb.sheetnames)
            finally:
                wb.close()
        finally:
            if process.poll() is None:
                process.terminate()
                process.wait(timeout=5)
            if process.stdout is not None:
                process.stdout.close()
            if process.stderr is not None:
                process.stderr.close()

    def test_vant_visual_contract_with_custom_action_buttons_is_preserved(self):
        html = (RESOLVER_DIR / "merge_ui.html").read_text(encoding="utf-8")
        required_fragments = [
            'id="fileTabs"',
            'id="sheetTabs"',
            'id="allLocalBtn"',
            'id="allRemoteBtn"',
            'id="cancelBtn"',
            "function cancelAndExit(",
            "fetch('/api/shutdown'",
            "function cycleCell(",
            "function bulkRow(",
            "function bulkCol(",
            "function showTip(",
            "className = 'luban-header'",
            "file.oursAuthor",
            "file.theirsAuthor",
            "file.oursDate",
            "file.theirsDate",
            "--bg-deep: #f4eed8",
            "--bg-root: #e6dfc8",
            "--gold: #f2c94c",
            "--lava: #f29aa3",
            "--emerald: #a2d2ff",
            "--sapphire: #9dccff",
            "--border-width: 3px",
            ".tooltip .val-local { color: #246aa3; }",
            ".tooltip .val-remote { color: #c0392b; }",
            "overflow-x: auto",
            "flex-wrap: nowrap",
            "function enableHorizontalWheel(",
            "enableHorizontalWheel(byId('fileTabs'))",
            "enableHorizontalWheel(byId('sheetTabs'))",
            "--incoming: #cfe9b7",
            "--tombstone: #77736b",
            ".axis-tombstone",
            ".axis-added-incoming",
            "function toggleAxis(",
            "function chooseRename(",
            "function bulkStructuralRow(",
            "function bulkStructuralColumn(",
            "target.rows",
            "target.columns",
            "主行列号＝合并后位置",
            "function structuralOutputMaps(",
            "function structuralCellSelectable(",
            ".cell-selectable",
            ".structure-badge.local-existing",
            "kept ? '已保留' : '已删除'",
            "合并后 ${output.columns.get(column.id) || '删除'}",
            "html { width: 100%; height: 100%; overflow: hidden; }",
            ".table-container {\n  flex: 1 1 auto;\n  min-width: 0;\n  min-height: 0;",
            ".footer {\n  position: relative;\n  z-index: 20;\n  flex: 0 0 auto;",
        ]
        for fragment in required_fragments:
            self.assertIn(fragment, html)

    def test_launch_returns_background_url_without_manual_diff_step(self):
        runtime = Path(self.temp.name) / "launch-runtime"
        entry = RESOLVER_DIR / "resolve_xlsx_conflict.py"
        completed = subprocess.run(
            [
                sys.executable,
                str(entry),
                "--repo",
                str(self.repo),
                "--runtime-dir",
                str(runtime),
                "launch",
                "--no-browser",
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=20,
        )
        self.assertEqual(completed.returncode, 0, completed.stdout + completed.stderr)
        result = json.loads(completed.stdout)
        self.assertTrue(result["ok"])
        self.assertTrue(result["url"].startswith("http://127.0.0.1:"))
        self.assertTrue(Path(result["manifest"]).is_file())

        with urllib.request.urlopen(f"{result['url']}/api/diff", timeout=5) as response:
            payload = json.load(response)
        self.assertEqual(payload["summary"]["fileCount"], 1)
        shutdown = urllib.request.Request(f"{result['url']}/api/shutdown", data=b"", method="POST")
        with urllib.request.urlopen(shutdown, timeout=5) as response:
            self.assertTrue(json.load(response)["success"])
        wait_for_process_exit(result["pid"])

    def test_cli_repo_defaults_to_current_directory(self):
        # 回归：--repo 缺省必须解析为当前目录（曾误为字面逗号 ","，导致不传 --repo 时失败）
        entry = RESOLVER_DIR / "resolve_xlsx_conflict.py"
        completed = subprocess.run(
            [sys.executable, str(entry), "detect"],
            cwd=str(self.repo),
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=30,
        )
        self.assertEqual(completed.returncode, 0, completed.stdout + completed.stderr)
        payload = json.loads(completed.stdout)
        self.assertTrue(payload["ok"])
        self.assertEqual(Path(payload["repo"]), self.repo.resolve())
        self.assertEqual(payload["count"], 1)

    def _full_decisions(self, manifest):
        """从 manifest 生成覆盖全部冲突 Cell 与 conflict_* Sheet 的决策 JSON。"""
        decisions = {"files": {}}
        for file_entry in manifest["files"]:
            sheets = {}
            for sheet in file_entry["diff"]["sheets"]:
                entry = {"cells": {}}
                for cell in sheet["cells"]:
                    if cell["resolution"] == "conflict":
                        entry["cells"][cell.get("id") or cell["coordinate"]] = "theirs"
                if sheet["action"].startswith("conflict_"):
                    entry["sheet"] = "theirs"
                if entry["cells"] or "sheet" in entry:
                    sheets[sheet["name"]] = entry
            if sheets:
                decisions["files"][file_entry["path"]] = {"sheets": sheets}
        return decisions

    def test_apply_works_from_non_git_cwd_via_manifest_repo_root(self):
        # serve/apply 不要求调用方 cwd 位于 git 仓库内：仓库身份来自 manifest.repoRoot
        runtime = Path(self.temp.name) / "runtime-apply"
        manifest_path = prepare_conflicts(self.repo, runtime_dir=runtime)
        decisions_path = Path(self.temp.name) / "decisions.json"
        decisions_path.write_text(
            json.dumps(self._full_decisions(load_manifest(manifest_path))),
            encoding="utf-8",
        )
        entry = RESOLVER_DIR / "resolve_xlsx_conflict.py"
        completed = subprocess.run(
            [sys.executable, str(entry), "apply",
             "--manifest", str(manifest_path), "--decisions", str(decisions_path), "--no-commit"],
            cwd=str(Path(self.temp.name)),  # 非 git 目录
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=30,
        )
        self.assertEqual(completed.returncode, 0, completed.stdout + completed.stderr)
        result = json.loads(completed.stdout)
        self.assertTrue(result["success"], result)
        wb = load_workbook(self.path, data_only=False)
        try:
            self.assertEqual(wb["Alpha"]["B2"].value, "theirs-alpha")
        finally:
            wb.close()

    def test_serve_works_from_non_git_cwd_and_exit_1_without_resolution(self):
        # serve 仅依赖 manifest；未应用解析时关闭服务器会以 exit 1 退出（0 仅代表已写回）
        runtime = Path(self.temp.name) / "runtime-serve"
        manifest_path = prepare_conflicts(self.repo, runtime_dir=runtime)
        entry = RESOLVER_DIR / "resolve_xlsx_conflict.py"
        process = subprocess.Popen(
            [sys.executable, str(entry), "serve", "--manifest", str(manifest_path), "--no-browser"],
            cwd=str(Path(self.temp.name)),  # 非 git 目录
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            encoding="utf-8",
        )
        try:
            line = process.stdout.readline().strip()
            self.assertTrue(line.startswith("MERGE_SERVER_URL="), line)
            url = line.split("=", 1)[1]
            with urllib.request.urlopen(f"{url}/api/diff", timeout=5) as response:
                payload = json.load(response)
            self.assertEqual(payload["summary"]["fileCount"], 1)
            shutdown = urllib.request.Request(f"{url}/api/shutdown", data=b"", method="POST")
            with urllib.request.urlopen(shutdown, timeout=5) as response:
                self.assertTrue(json.load(response)["success"])
            self.assertEqual(process.wait(timeout=10), 1)
        finally:
            if process.poll() is None:
                process.terminate()
                process.wait(timeout=5)
            if process.stdout is not None:
                process.stdout.close()
            if process.stderr is not None:
                process.stderr.close()

    def test_agents_doc_declares_ui_first_detect_then_launch_route(self):
        agent_doc = (PACKAGE_ROOT.parent / "AGENTS.md").read_text(encoding="utf-8")
        self.assertLess(agent_doc.find("--repo <repo> detect"), agent_doc.find("--repo <repo> launch"))
        self.assertIn("不调用通用 `xlsx` Skill", agent_doc)
        self.assertIn("不运行 `prepare` 后读取 manifest", agent_doc)
        self.assertIn("不使用阻塞式 `resolve` 作为正常入口", agent_doc)
        self.assertIn("不默认走无头自动合并", agent_doc)
        self.assertIn("默认流程必须走 UI `launch`", agent_doc)


class XlsmConflictTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.repo = Path(self.temp.name) / "meta"
        self.repo.mkdir()
        git(self.repo, "init")
        git(self.repo, "config", "user.name", "Macro Test")
        git(self.repo, "config", "user.email", "macro@example.invalid")
        self.path = self.repo / "Enums.xlsm"

    def tearDown(self):
        self.temp.cleanup()

    def create_conflict(self, base_macro, ours_macro, theirs_macro, *, include_xlsx=False):
        write_macro_book(self.path, "base", base_macro)
        other = self.repo / "Other.xlsx"
        if include_xlsx:
            write_grid(other, [["base"]])
        git(self.repo, "add", ".")
        git(self.repo, "commit", "-m", "base")
        git(self.repo, "checkout", "-b", "theirs")
        write_macro_book(self.path, "theirs", theirs_macro)
        if include_xlsx:
            write_grid(other, [["theirs"]])
        git(self.repo, "add", ".")
        git(self.repo, "commit", "-m", "theirs")
        git(self.repo, "checkout", "-")
        write_macro_book(self.path, "ours", ours_macro)
        if include_xlsx:
            write_grid(other, [["ours"]])
        git(self.repo, "add", ".")
        git(self.repo, "commit", "-m", "ours")
        self.assertNotEqual(git(self.repo, "merge", "theirs", check=False).returncode, 0)

    def test_xlsm_conflict_preserves_macro_payload_and_extension(self):
        macro = b"shared-vba-project"
        self.create_conflict(macro, macro, macro)
        self.assertEqual(detect_conflicts(self.repo), [{"path": "Enums.xlsm", "stages": [1, 2, 3]}])
        manifest_path = prepare_conflicts(self.repo, runtime_dir=Path(self.temp.name) / "runtime")
        manifest = load_manifest(manifest_path)
        file_entry = manifest["files"][0]
        self.assertTrue(file_entry["ours"].endswith("ours.xlsm"))
        self.assertEqual(file_entry["macro"]["oursEntries"], ["xl/vbaProject.bin"])
        self.assertEqual(file_entry["macro"]["status"], "same")
        sheet = file_entry["diff"]["sheets"][0]
        cell = next(item for item in sheet["cells"] if item["resolution"] == "conflict")
        cell_key = cell.get("id") or cell["coordinate"]
        result = apply_manifest(manifest_path, {
            "files": {"Enums.xlsm": {"sheets": {sheet["name"]: {"cells": {cell_key: "theirs"}}}}}
        }, commit=False)
        self.assertTrue(result["success"], result)
        wb = load_workbook(self.path, keep_vba=True, data_only=False)
        try:
            self.assertEqual(wb["Data"]["A1"].value, "theirs")
        finally:
            _close_workbook(wb)
        self.assertEqual(read_macro_payload(self.path), macro)

    def test_xlsm_conflict_exposes_divergent_macro_payload_choice(self):
        self.create_conflict(b"base-vba", b"local-vba", b"remote-vba")
        manifest_path = prepare_conflicts(self.repo, runtime_dir=Path(self.temp.name) / "runtime")
        file_entry = load_manifest(manifest_path)["files"][0]
        self.assertEqual(file_entry["macro"]["status"], "conflict")
        self.assertEqual(file_entry["macro"]["default"], "ours")
        sheet = file_entry["diff"]["sheets"][0]
        cell = next(item for item in sheet["cells"] if item["resolution"] == "conflict")
        key = cell.get("id") or cell["coordinate"]
        for side, expected in (("ours", b"local-vba"), ("theirs", b"remote-vba")):
            merged = _build_merged_file(file_entry, {
                "files": {"Enums.xlsm": {"macro": side, "sheets": {sheet["name"]: {"cells": {key: side}}}}}
            })
            try:
                self.assertEqual(read_macro_payload(merged), expected)
            finally:
                merged.unlink(missing_ok=True)

    def test_batch_prepare_keeps_xlsm_macro_conflict_with_other_workbooks(self):
        self.create_conflict(b"base-vba", b"local-vba", b"remote-vba", include_xlsx=True)
        manifest = load_manifest(prepare_conflicts(self.repo, runtime_dir=Path(self.temp.name) / "runtime"))
        self.assertEqual(manifest["summary"]["fileCount"], 2)
        self.assertEqual({item["path"] for item in manifest["files"]}, {"Enums.xlsm", "Other.xlsx"})


class StructuralMergeTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.base = self.root / "base.xlsx"
        self.ours = self.root / "ours.xlsx"
        self.theirs = self.root / "theirs.xlsx"

    def tearDown(self):
        self.temp.cleanup()

    def merge(self, diff, sheet_choices):
        file_entry = {
            "id": "Data.xlsx",
            "path": "Data.xlsx",
            "base": str(self.base),
            "ours": str(self.ours),
            "theirs": str(self.theirs),
            "output": str(self.root / "output.xlsx"),
            "diff": diff,
        }
        decisions = {"files": {"Data.xlsx": {"sheets": {"Data": sheet_choices}}}}
        return _build_merged_file(file_entry, decisions)

    def test_same_only_changes_sheet_is_folded_away(self):
        # 双方一致变化（resolution=same）：无决策项 → needsAttention=False 且行折叠
        rows_base = [[1, "a", "x"], [2, "b", "y"], [3, "c", "z"]]
        rows_changed = [[1, "a", "x"], [2, "B", "y"], [3, "c", "z"]]
        write_named_grid(self.base, "Data", rows_base)
        write_named_grid(self.ours, "Data", rows_changed)
        write_named_grid(self.theirs, "Data", rows_changed)
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        self.assertFalse(sheet["needsAttention"])
        self.assertEqual(sheet["rows"], [])
        self.assertEqual(len(sheet["allRows"]), 3)
        self.assertEqual(sheet["conflictCount"], 0)

    def test_luban_anonymous_comment_column_survives_structural_rebuild(self):
        write_luban_anonymous_book(self.base, "base")
        write_luban_anonymous_book(self.ours, "local")
        write_luban_anonymous_book(self.theirs, "remote")
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        anonymous = next(column for column in sheet["columns"] if column.get("baseColumn") == "C")
        self.assertEqual(anonymous["id"], "@3")
        conflict = next(cell for cell in sheet["cells"] if cell["resolution"] == "conflict")
        merged = self.merge(diff, {"cells": {conflict["id"]: "theirs"}})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                ws = wb["Data"]
                self.assertEqual(ws.max_column, 4)
                self.assertIsNone(ws["C1"].value)
                self.assertEqual(ws["C4"].value, "策划注释")
                self.assertEqual(ws["C5"].value, "行军队列")
                self.assertEqual(ws["D1"].value, "value")
                self.assertEqual(ws["D5"].value, "remote")
                self.assertEqual(ws.column_dimensions["C"].width, 19.25)
                self.assertEqual(ws["C5"].fill.fgColor.rgb, "00FFF2CC")
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_header_row_heights_survive_structural_rebuild(self):
        # 表头行高在结构重建后必须保留（meta 仓库大量表有自定义行高）
        for path, value in ((self.base, "base"), (self.ours, "local"), (self.theirs, "remote")):
            write_luban_anonymous_book(path, value)
            wb = load_workbook(path)
            ws = wb["Data"]
            ws.row_dimensions[1].height = 30.0   # ##var
            ws.row_dimensions[2].height = 45.0   # ##type
            ws.row_dimensions[4].height = 60.0   # ##
            wb.save(path)
            wb.close()
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        conflict = next(cell for cell in sheet["cells"] if cell["resolution"] == "conflict")
        merged = self.merge(diff, {"cells": {conflict["id"]: "ours"}})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                ws = wb["Data"]
                self.assertEqual(ws.row_dimensions[1].height, 30.0)
                self.assertEqual(ws.row_dimensions[2].height, 45.0)
                self.assertEqual(ws.row_dimensions[4].height, 60.0)
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_formula_cell_cache_survives_structural_rebuild(self):
        # 公式 cell 重建后必须保留公式和缓存值：Luban 用 data_only 读取缓存值
        import zipfile
        import xml.etree.ElementTree as ET

        def save_with_cache(path, c5, cached_b5):
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["##var", "id", "value"])
            ws.append(["##type", "int", "int"])
            ws.append(["##group", None, "c"])
            ws.append(["##", "ID", "值"])
            ws.append([None, 1, c5])
            ws["B5"] = "=C5*10"
            wb.save(path)
            wb.close()
            # 手动注入缓存值（模拟 Excel 打开后的状态）
            with zipfile.ZipFile(path) as z:
                infos = z.infolist()
                parts = {i.filename: z.read(i.filename) for i in infos}
            tree = ET.fromstring(parts['xl/worksheets/sheet1.xml'])
            NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
            for c in tree.iter(f'{NS}c'):
                if c.get('r') == 'B5':
                    v = c.find(f'{NS}v')
                    if v is None:
                        v = ET.SubElement(c, f'{NS}v')
                    v.text = str(cached_b5)
            parts['xl/worksheets/sheet1.xml'] = ET.tostring(tree, encoding='UTF-8')
            with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as z:
                for info in infos:
                    z.writestr(info, parts[info.filename])

        save_with_cache(self.base, 100, 1000)
        save_with_cache(self.ours, 150, 1500)
        save_with_cache(self.theirs, 200, 2000)

        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        # 公式相同但引用值不同 -> 缓存值不同 -> conflict
        conflict = next(cell for cell in sheet["cells"] if cell["resolution"] == "conflict")
        self.assertEqual(conflict["ours"], 1500)
        self.assertEqual(conflict["theirs"], 2000)
        merged = self.merge(diff, {"cells": {conflict["id"]: "ours"}})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                self.assertEqual(wb["Data"]["B5"].value, "=C5*10")
            finally:
                wb.close()
            wb_v = load_workbook(merged, data_only=True)
            try:
                self.assertEqual(wb_v["Data"]["B5"].value, 1500)
            finally:
                wb_v.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_noncanonical_freeze_pane_view_survives_structural_rebuild(self):
        write_luban_anonymous_book(self.base, "base", pane_top_left="D32")
        write_luban_anonymous_book(self.ours, "local", pane_top_left="D32")
        write_luban_anonymous_book(self.theirs, "remote", pane_top_left="D32")
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        conflict = next(cell for cell in sheet["cells"] if cell["resolution"] == "conflict")
        merged = self.merge(diff, {"cells": {conflict["id"]: "ours"}})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                pane = wb["Data"].sheet_view.pane
                self.assertEqual(pane.state, "frozen")
                self.assertEqual(pane.topLeftCell, "D32")
                self.assertEqual(pane.xSplit, 3)
                self.assertEqual(pane.ySplit, 4)
                self.assertEqual(pane.activePane, "bottomRight")
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_noncanonical_freeze_pane_view_survives_added_sheet_copy(self):
        for path in (self.base, self.ours, self.theirs):
            wb = Workbook()
            wb.active.title = "Data"
            wb["Data"]["A1"] = "same"
            if path == self.theirs:
                extra = wb.create_sheet("RemoteSheet")
                extra["A1"] = "remote"
                extra.sheet_view.pane = Pane(
                    xSplit=3, ySplit=4, topLeftCell="D32",
                    activePane="bottomRight", state="frozen",
                )
            wb.save(path)
            wb.close()
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        merged = self.merge(diff, {})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                pane = wb["RemoteSheet"].sheet_view.pane
                self.assertEqual((pane.topLeftCell, pane.xSplit, pane.ySplit), ("D32", 3, 4))
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_merged_header_range_survives_structural_rebuild(self):
        write_structured_header_book(self.base, "base")
        write_structured_header_book(self.ours, "local")
        write_structured_header_book(self.theirs, "remote")
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        conflict = next(cell for cell in sheet["cells"] if cell["resolution"] == "conflict")
        merged = self.merge(diff, {"cells": {conflict["id"]: "theirs"}})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                self.assertEqual({str(item) for item in wb["Data"].merged_cells.ranges}, {"I1:M1"})
                self.assertEqual(wb["Data"]["J2"].value, "alias")
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_unchanged_auxiliary_sheet_is_preserved_without_rebuild(self):
        write_structured_header_book(self.base, "base", include_aux=True)
        write_structured_header_book(self.ours, "local", include_aux=True)
        write_structured_header_book(self.theirs, "remote", include_aux=True)
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        data = next(sheet for sheet in diff["sheets"] if sheet["name"] == "Data")
        conflict = next(cell for cell in data["cells"] if cell["resolution"] == "conflict")
        merged = self.merge(diff, {"cells": {conflict["id"]: "ours"}})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                aux = wb["Aux"]
                self.assertEqual(aux.max_column, 16)
                self.assertEqual(aux["C1"].value, "right")
                self.assertEqual({str(item) for item in aux.merged_cells.ranges}, {"E1:G1"})
                self.assertEqual(aux["P10"].fill.fgColor.rgb, "00FFF2CC")
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_divergent_merged_header_changes_require_and_apply_structure_choice(self):
        write_structured_header_book(self.base, "base", merged_range="I1:M1")
        write_structured_header_book(self.ours, "local", merged_range="I1:L1")
        write_structured_header_book(self.theirs, "remote", merged_range="I1:N1")
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        self.assertEqual(sheet["mergedCells"]["status"], "conflict")
        self.assertTrue(sheet["needsAttention"])
        conflict = next(cell for cell in sheet["cells"] if cell["resolution"] == "conflict")
        for side, expected in (("ours", "I1:L1"), ("theirs", "I1:N1")):
            merged = self.merge(diff, {
                "mergedCells": side,
                "cells": {conflict["id"]: side},
            })
            try:
                wb = load_workbook(merged, data_only=False)
                try:
                    self.assertEqual({str(item) for item in wb["Data"].merged_cells.ranges}, {expected})
                finally:
                    wb.close()
            finally:
                merged.unlink(missing_ok=True)

    def test_one_sided_merged_header_change_is_applied_automatically(self):
        write_structured_header_book(self.base, "base", merged_range="I1:M1")
        write_structured_header_book(self.ours, "base", merged_range="I1:M1")
        write_structured_header_book(self.theirs, "remote", merged_range="I1:N1")
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        self.assertEqual(sheet["mergedCells"]["status"], "auto_theirs")
        merged = self.merge(diff, {})
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                self.assertEqual({str(item) for item in wb["Data"].merged_cells.ranges}, {"I1:N1"})
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_conflict_sheet_keeps_attention_and_conflict_rows(self):
        # 同一 cell 三方各不同：needsAttention=True，冲突行保留展示
        rows_base = [[1, "a", "x"], [2, "b", "y"], [3, "c", "z"]]
        write_named_grid(self.base, "Data", rows_base)
        write_named_grid(self.ours, "Data", [[1, "a", "x"], [2, "B", "y"], [3, "c", "z"]])
        write_named_grid(self.theirs, "Data", [[1, "a", "x"], [2, "C", "y"], [3, "c", "z"]])
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = diff["sheets"][0]
        self.assertTrue(sheet["needsAttention"])
        self.assertEqual(sheet["conflictCount"], 1)
        self.assertEqual(len(sheet["rows"]), 1)  # 只有冲突行（same 行折叠）
        self.assertEqual(len(sheet["allRows"]), 3)

    def test_hidden_sheet_state_is_preserved_after_merge(self):
        # 隐藏页签状态在写回重建后必须保留
        for path in (self.base, self.ours, self.theirs):
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["##var", "value"])
            ws.append(["id", "name"])
            ws.append([1, "a"])
            hidden = wb.create_sheet("Hidden")
            hidden["A1"] = "x"
            hidden.sheet_state = "hidden"
            wb.save(path)
            wb.close()
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        # 隐藏页签默认不参与冲突检测
        self.assertEqual([sh["name"] for sh in diff["sheets"]], ["Data"])
        out = self.merge(diff, {})
        wb = load_workbook(out)
        self.assertEqual(wb["Data"].sheet_state, "visible")
        self.assertEqual(wb["Hidden"].sheet_state, "hidden")
        wb.close()
        # 显式开启后可参与检测
        diff_all = diff_workbooks(
            self.base, self.ours, self.theirs,
            analysis_dir=self.root / "analysis2", skip_hidden_sheets=False,
        )
        self.assertIn("Hidden", [sh["name"] for sh in diff_all["sheets"]])

    def test_incoming_row_and_column_are_locked_and_rebuilt_with_cell_choice(self):
        write_grid(self.base, [[1, 10, 100], [2, 20, 200], [3, 30, 300]])
        write_grid(self.ours, [[1, 11, 100], [2, 20, 200], [3, 30, 300]])
        write_grid(
            self.theirs,
            [[1, 100, "remote-a"], [4, 400, "remote-new"], [2, 200, "remote-b"], [3, 300, "remote-c"]],
        )
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = next(item for item in diff["sheets"] if item["name"] == "Data")

        incoming_rows = [row for row in sheet["allRows"] if row["state"] == "added" and row["incoming"]]
        incoming_columns = [column for column in sheet["columns"] if column["state"] == "added" and column["incoming"]]
        delete_modify_columns = [column for column in sheet["columns"] if column["state"] == "delete_modify"]
        self.assertEqual(len(incoming_rows), 1)
        self.assertEqual(len(incoming_columns), 1)
        self.assertEqual(len(delete_modify_columns), 1)
        self.assertTrue(incoming_rows[0]["locked"])
        self.assertTrue(incoming_columns[0]["locked"])

        column = delete_modify_columns[0]
        conflict = next(cell for cell in sheet["cells"] if cell["resolution"] == "conflict")
        merged = self.merge(diff, {
            "rows": {},
            "columns": {column["id"]: "keep"},
            "cells": {conflict["id"]: "ours"},
        })
        try:
            wb = load_workbook(merged, data_only=False)
            try:
                values = list(wb["Data"].values)
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)
        self.assertEqual(values, [
            (1, 11, 100, "remote-a"),
            (4, None, 400, "remote-new"),
            (2, 20, 200, "remote-b"),
            (3, 30, 300, "remote-c"),
        ])

    def test_delete_modify_row_can_be_kept_or_rendered_as_a_tombstone(self):
        write_grid(self.base, [[1, "a"], [2, "b"], [3, "c"]])
        write_grid(self.ours, [[1, "a"], [2, "local"], [3, "c"]])
        write_grid(self.theirs, [[1, "a"], [3, "c"]])
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = next(item for item in diff["sheets"] if item["name"] == "Data")
        row = next(item for item in sheet["allRows"] if item["state"] == "delete_modify")
        self.assertFalse(row["locked"])
        self.assertEqual(row["default"], "keep")
        conflict_cells = [cell for cell in sheet["cells"] if cell["rowId"] == row["id"] and cell["resolution"] == "conflict"]

        keep_choices = {cell["id"]: "ours" for cell in conflict_cells}
        kept = self.merge(diff, {"rows": {row["id"]: "keep"}, "columns": {}, "cells": keep_choices})
        try:
            wb = load_workbook(kept)
            try:
                self.assertEqual(list(wb["Data"].values), [(1, "a"), (2, "local"), (3, "c")])
            finally:
                wb.close()
        finally:
            kept.unlink(missing_ok=True)

        deleted = self.merge(diff, {"rows": {row["id"]: "delete"}, "columns": {}, "cells": keep_choices})
        try:
            wb = load_workbook(deleted)
            try:
                self.assertEqual(list(wb["Data"].values), [(1, "a"), (3, "c")])
            finally:
                wb.close()
        finally:
            deleted.unlink(missing_ok=True)

    def test_sheet_rename_conflict_uses_the_selected_name(self):
        rows = [[1, "a"], [2, "b"]]
        write_named_grid(self.base, "BaseName", rows)
        write_named_grid(self.ours, "OursName", rows)
        write_named_grid(self.theirs, "TheirsName", rows)
        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = next(item for item in diff["sheets"] if item["renameStatus"] == "rename_conflict")
        file_entry = {
            "id": "Data.xlsx",
            "path": "Data.xlsx",
            "base": str(self.base),
            "ours": str(self.ours),
            "theirs": str(self.theirs),
            "output": str(self.root / "output.xlsx"),
            "diff": diff,
        }
        decisions = {
            "files": {
                "Data.xlsx": {
                    "sheets": {
                        sheet["name"]: {"rename": "theirs", "rows": {}, "columns": {}, "cells": {}}
                    }
                }
            }
        }
        merged = _build_merged_file(file_entry, decisions)
        try:
            wb = load_workbook(merged)
            try:
                self.assertEqual(wb.sheetnames, ["TheirsName"])
                self.assertEqual(list(wb["TheirsName"].values), [(1, "a"), (2, "b")])
            finally:
                wb.close()
        finally:
            merged.unlink(missing_ok=True)

    def test_realistic_repeated_grid_does_not_create_shifted_fake_diffs(self):
        standard = [1, 2, 3, 5, 6, 7, 0, 8, 9]
        write_grid(self.base, [standard, standard, standard, [0] * 9, *([standard] * 6), [9] * 9])
        write_grid(self.ours, [
            [1, 2, "?", 5, 5, 6, 7, 8, 9],
            *([[1, 2, "?", 4, 5, 6, 7, 8, 9]] * 4),
            *([[1, 2, 3, 4, 5, 6, 7, 8, 9]] * 4),
            [10] * 9,
        ])
        theirs = [[*row[:2], "#", *row[3:], None, None] for row in [
            standard, standard, standard, [0] * 9, *([standard] * 6), [9] * 9
        ]]
        theirs.append([None] * 10 + [10])
        write_grid(self.theirs, theirs)

        diff = diff_workbooks(self.base, self.ours, self.theirs, analysis_dir=self.root / "analysis")
        sheet = next(item for item in diff["sheets"] if item["name"] == "Data")
        self.assertEqual(
            [(column["status"], column.get("baseColumn"), column.get("oursColumn"), column.get("theirsColumn"))
             for column in sheet["columns"]],
            [
                ("common", "A", "A", "A"),
                ("common", "B", "B", "B"),
                ("common", "C", "C", "C"),
                ("added_ours", None, "D", None),
                ("common", "D", "E", "D"),
                ("common", "E", "F", "E"),
                ("common", "F", "G", "F"),
                ("deleted_ours", "G", None, "G"),
                ("common", "H", "H", "H"),
                ("common", "I", "I", "I"),
                ("added_theirs", None, None, "J"),
                ("added_theirs", None, None, "K"),
            ],
        )
        self.assertEqual(
            [(row["status"], row.get("baseRow"), row.get("oursRow"), row.get("theirsRow"))
             for row in sheet["allRows"]],
            [
                ("common", 1, 1, 1),
                ("common", 2, 2, 2),
                ("common", 3, 3, 3),
                ("delete_modify_ours", 4, None, 4),
                ("common", 5, 4, 5),
                ("common", 6, 5, 6),
                ("common", 7, 6, 7),
                ("common", 8, 7, 8),
                ("common", 9, 8, 9),
                ("common", 10, 9, 10),
                ("delete_modify_ours", 11, None, 11),
                ("added_ours", None, 10, None),
                ("added_theirs", None, None, 12),
            ],
        )
        self.assertEqual(sheet["changedCount"], sheet["decisionCellCount"])
        self.assertLess(sheet["changedCount"], sheet["rawChangedCellCount"])


if __name__ == "__main__":
    unittest.main()
