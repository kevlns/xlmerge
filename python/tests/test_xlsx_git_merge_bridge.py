import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PACKAGE_ROOT))

from xlsx_merge_engine import xlsx_git_merge_bridge as bridge_mod
from xlsx_merge_engine.xlsx_git_merge_bridge import RowProjection, analyze_workbooks, _row_side_alignment


def write_book(path, fields, rows, *, sheet="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["##var", *fields])
    ws.append(["##type", *(["string"] * len(fields))])
    ws.append(["##", *([None] * len(fields))])
    for row in rows:
        ws.append([None, *(row.get(field) for field in fields)])
    wb.save(path)
    wb.close()


def write_raw_book(path, rows, *, sheet="Sheet1"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


def write_raw_sheets(path, sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
    wb.save(path)
    wb.close()


def write_luban_anonymous_book(path, rows, *, include_anonymous=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    if include_anonymous:
        ws.append(["##var", "id", None, "value"])
        ws.append(["##type", "int", None, "string"])
        ws.append(["##group", None, None, "c"])
        ws.append(["##", "ID", "策划注释", "值"])
        for row_id, note, value in rows:
            ws.append([None, row_id, note, value])
    else:
        ws.append(["##var", "id", "value"])
        ws.append(["##type", "int", "string"])
        ws.append(["##group", None, "c"])
        ws.append(["##", "ID", "值"])
        for row_id, _note, value in rows:
            ws.append([None, row_id, value])
    wb.save(path)
    wb.close()


def merged_rows(sheet_report):
    rows = []
    for segment in sheet_report["segments"]:
        if segment["type"] in {"merged", "resolved"}:
            rows.extend(segment["rows"])
    return rows


class XlsxGitMergeBridgeTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.root = Path(self.temp.name)
        self.base = self.root / "base.xlsx"
        self.ours = self.root / "ours.xlsx"
        self.theirs = self.root / "theirs.xlsx"
        self.output = self.root / "output"

    def tearDown(self):
        self.temp.cleanup()

    def test_luban_anonymous_comment_column_is_aligned(self):
        rows = [(1, "alpha note", "a"), (2, "beta note", "b")]
        write_luban_anonymous_book(self.base, rows)
        write_luban_anonymous_book(self.ours, [(1, "alpha note", "local"), rows[1]])
        write_luban_anonymous_book(self.theirs, [(1, "remote note", "a"), rows[1]])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        anonymous = next(column for column in sheet["columnAlignment"] if column["baseColumn"] == "C")
        self.assertEqual(anonymous["field"], "@3")
        self.assertEqual(anonymous["status"], "common")
        self.assertEqual((anonymous["oursColumn"], anonymous["theirsColumn"]), ("C", "C"))
        self.assertIn("@3", sheet["projectionFields"])

    def test_luban_anonymous_delete_modify_is_a_column_conflict(self):
        rows = [(1, "alpha note", "a"), (2, "beta note", "b")]
        write_luban_anonymous_book(self.base, rows)
        write_luban_anonymous_book(self.ours, rows, include_anonymous=False)
        write_luban_anonymous_book(self.theirs, [(1, "changed note", "a"), rows[1]])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        anonymous = next(column for column in sheet["columnAlignment"] if column["baseColumn"] == "C")
        self.assertEqual(anonymous["status"], "delete_modify_ours")
        self.assertTrue(anonymous["conflict"])

    def test_generic_internal_blank_column_keeps_its_physical_position(self):
        rows = [["left", None, "right"], ["a", None, "b"]]
        write_raw_book(self.base, rows)
        write_raw_book(self.ours, [["left-local", None, "right"], rows[1]])
        write_raw_book(self.theirs, [rows[0], ["a", None, "b-remote"]])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(
            [(item["baseColumn"], item["oursColumn"], item["theirsColumn"])
             for item in sheet["columnAlignment"]],
            [("A", "A", "A"), ("B", "B", "B"), ("C", "C", "C")],
        )
        self.assertIn("@2", sheet["projectionFields"])

    def test_remote_row_insert_and_local_later_edit_merge_cleanly(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "c", "value": 30},
            {"id": 3, "name": "d", "value": 40},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "c", "value": 30},
            {"id": 3, "name": "d", "value": 41},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
            {"id": 3, "name": "c", "value": 30},
            {"id": 4, "name": "d", "value": 40},
        ])

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(sheet["projectionFields"], ["name", "value"])
        self.assertEqual(merged_rows(sheet), [["a", 10], ["b", 20], ["c", 30], ["d", 41]])

    def test_remote_multi_row_insert_and_local_later_edit_merge_cleanly(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "c", "value": 30},
            {"id": 3, "name": "d", "value": 40},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "c", "value": 30},
            {"id": 3, "name": "d", "value": 41},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b1", "value": 20},
            {"id": 3, "name": "b2", "value": 21},
            {"id": 4, "name": "c", "value": 30},
            {"id": 5, "name": "d", "value": 40},
        ])

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(
            merged_rows(sheet),
            [["a", 10], ["b1", 20], ["b2", 21], ["c", 30], ["d", 41]],
        )

    def test_single_row_delete_merges_cleanly(self):
        fields = ["id", "name"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a"},
            {"id": 2, "name": "b"},
            {"id": 3, "name": "c"},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a"},
            {"id": 3, "name": "c"},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a"},
            {"id": 2, "name": "b"},
            {"id": 3, "name": "c"},
        ])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        self.assertEqual(report["sheets"][0]["status"], "clean")
        self.assertEqual(merged_rows(report["sheets"][0]), [[1, "a"], [3, "c"]])

    def test_row_insert_and_different_row_delete_merge_cleanly(self):
        fields = ["id", "name"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a"},
            {"id": 2, "name": "b"},
            {"id": 3, "name": "c"},
            {"id": 4, "name": "d"},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a"},
            {"id": 5, "name": "new"},
            {"id": 2, "name": "b"},
            {"id": 3, "name": "c"},
            {"id": 4, "name": "d"},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a"},
            {"id": 2, "name": "b"},
            {"id": 4, "name": "d"},
        ])

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        self.assertEqual(report["sheets"][0]["status"], "clean")
        self.assertEqual(
            merged_rows(report["sheets"][0]),
            [["a"], ["new"], ["b"], ["d"]],
        )

    def test_disjoint_change_segments_use_surrounding_anchors(self):
        fields = ["id", "name", "value"]
        base_rows = [
            {"id": index, "name": name, "value": index * 10}
            for index, name in enumerate("abcdef", 1)
        ]
        ours_rows = [base_rows[0], {"id": 7, "name": "x", "value": 70}, *base_rows[1:4],
                     {"id": 5, "name": "e", "value": 51}, base_rows[5]]
        theirs_rows = [*base_rows[:3], {"id": 8, "name": "y", "value": 80}, *base_rows[3:5]]
        write_book(self.base, fields, base_rows)
        write_book(self.ours, fields, ours_rows)
        write_book(self.theirs, fields, theirs_rows)

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(
            merged_rows(sheet),
            [["a", 10], ["x", 70], ["b", 20], ["c", 30], ["y", 80],
             ["d", 40], ["e", 51]],
        )

    def test_same_row_different_cell_change_produces_cell_conflict(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [{"id": 1, "name": "a", "value": 10}])
        write_book(self.ours, fields, [{"id": 1, "name": "a", "value": 11}])
        write_book(self.theirs, fields, [{"id": 1, "name": "a", "value": 12}])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "conflict")
        conflict = next(segment for segment in sheet["segments"] if segment["type"] == "conflict")
        self.assertFalse(conflict["alignmentRequired"])
        value_diff = next(
            cell for cell in conflict["rowSlots"][0]["cellDiffs"] if cell["field"] == "value"
        )
        self.assertEqual(value_diff["base"], 10)
        self.assertEqual(value_diff["ours"], 11)
        self.assertEqual(value_diff["theirs"], 12)
        self.assertEqual(value_diff["resolution"], "conflict")

    def test_same_row_different_columns_merge_at_cell_level(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [{"id": 1, "name": "a", "value": 10}])
        write_book(self.ours, fields, [{"id": 1, "name": "local", "value": 10}])
        write_book(self.theirs, fields, [{"id": 1, "name": "a", "value": 12}])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(sheet["gitRawStatus"], "conflict")
        self.assertEqual(merged_rows(sheet), [[1, "local", 12]])

    def test_both_insert_different_rows_at_same_anchor_preserves_local_then_remote(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "c", "value": 30},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "ours-new", "value": 20},
            {"id": 3, "name": "c", "value": 30},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "theirs-new", "value": 21},
            {"id": 3, "name": "c", "value": 30},
        ])

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        resolved = next(
            segment for segment in sheet["segments"] if segment["type"] == "resolved"
        )
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(resolved["policy"], "same_anchor_insert_union_ours_first")
        self.assertEqual(
            merged_rows(sheet),
            [["a", 10], ["ours-new", 20], ["theirs-new", 21], ["c", 30]],
        )

    def test_both_insert_identical_rows_at_same_anchor_collapses_duplicate(self):
        fields = ["id", "name", "value"]
        base_rows = [
            {"id": 1, "name": "a", "value": 10},
            {"id": 3, "name": "c", "value": 30},
        ]
        inserted_rows = [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
            {"id": 3, "name": "c", "value": 30},
        ]
        write_book(self.base, fields, base_rows)
        write_book(self.ours, fields, inserted_rows)
        write_book(self.theirs, fields, inserted_rows)

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(merged_rows(sheet), [["a", 10], ["b", 20], ["c", 30]])

    def test_delete_vs_modify_is_reported_as_row_level_candidate(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
            {"id": 3, "name": "c", "value": 30},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 3, "name": "c", "value": 30},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 22},
            {"id": 3, "name": "c", "value": 30},
        ])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        conflict = next(
            segment for segment in report["sheets"][0]["segments"]
            if segment["type"] == "conflict"
        )
        self.assertTrue(conflict["alignmentRequired"])
        self.assertEqual(conflict["rowSlots"][0]["kind"], "delete_modify_ours")

    def test_remote_added_column_is_excluded_from_row_projection_and_reported(self):
        base_fields = ["id", "name", "value"]
        theirs_fields = ["id", "name", "bonus", "value"]
        write_book(self.base, base_fields, [{"id": 1, "name": "a", "value": 10}])
        write_book(self.ours, base_fields, [{"id": 1, "name": "a", "value": 11}])
        write_book(self.theirs, theirs_fields, [
            {"id": 1, "name": "a", "bonus": "new", "value": 10}
        ])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(sheet["projectionFields"], ["id", "name", "value"])
        self.assertEqual(sheet["columns"]["theirsOnly"], ["bonus"])
        self.assertEqual(merged_rows(sheet), [[1, "a", 11]])

        sidecar = json.loads(
            (Path(sheet["artifacts"]["directory"]) / "theirs.sidecar.json").read_text(encoding="utf-8")
        )
        self.assertEqual(sidecar["rows"][0]["cells"]["bonus"], "new")

    def test_deleted_column_conflicts_with_other_side_cell_modification(self):
        base_fields = ["id", "name", "value"]
        write_book(self.base, base_fields, [{"id": 1, "name": "a", "value": 10}])
        write_book(self.ours, base_fields, [{"id": 1, "name": "a", "value": 11}])
        write_book(self.theirs, ["id", "name"], [{"id": 1, "name": "a"}])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        column = next(item for item in sheet["columnAlignment"] if item["field"] == "value")
        self.assertEqual(sheet["gitStatus"], "clean")
        self.assertEqual(sheet["status"], "conflict")
        self.assertEqual(sheet["columnConflictCount"], 1)
        self.assertEqual(report["summary"]["columnConflictCount"], 1)
        self.assertEqual(column["status"], "delete_modify_theirs")
        self.assertTrue(column["conflict"])
        self.assertEqual(column["contentChanges"], [{
            "kind": "modified_ours",
            "baseXlsxRow": 4,
            "oursXlsxRow": 4,
            "base": 10,
            "ours": 11,
        }])

    def test_deleted_unchanged_column_is_not_a_structural_conflict(self):
        base_fields = ["id", "name", "value"]
        rows = [{"id": 1, "name": "a", "value": 10}]
        write_book(self.base, base_fields, rows)
        write_book(self.ours, base_fields, rows)
        write_book(self.theirs, ["id", "name"], [{"id": 1, "name": "a"}])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        column = next(item for item in sheet["columnAlignment"] if item["field"] == "value")
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(sheet["columnConflictCount"], 0)
        self.assertEqual(column["status"], "deleted_theirs")
        self.assertNotIn("contentChanges", column)

    def test_deleted_column_conflicts_with_inserted_row_value(self):
        base_fields = ["id", "name", "value"]
        write_book(self.base, base_fields, [{"id": 1, "name": "a", "value": 10}])
        write_book(self.ours, base_fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
        ])
        write_book(self.theirs, ["id", "name"], [{"id": 1, "name": "a"}])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        column = next(
            item for item in report["sheets"][0]["columnAlignment"]
            if item["field"] == "value"
        )
        self.assertEqual(column["status"], "delete_modify_theirs")
        self.assertEqual(column["contentChanges"], [{
            "kind": "inserted_ours",
            "baseXlsxRow": None,
            "oursXlsxRow": 5,
            "base": None,
            "ours": 20,
        }])

    def test_headerless_middle_column_and_row_insert_do_not_expand_row_conflict(self):
        original = list(range(1, 10))
        base_rows = [original[:] for _ in range(10)]
        ours_rows = [original[:] for _ in range(9)] + [[10] * 9]
        theirs_rows = []
        for index, row in enumerate(base_rows):
            values = row[:7] + [0] + row[7:]
            if index == 3:
                theirs_rows.append([0] * 10)
            theirs_rows.append(values)
        theirs_rows[-1] = [9] * 7 + [0, 9, 9]
        write_raw_book(self.base, base_rows)
        write_raw_book(self.ours, ours_rows)
        write_raw_book(self.theirs, theirs_rows)

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["columnAlignmentMode"], "content_anchor_sequence")
        self.assertEqual(sheet["projectionFields"], [f"@{index}" for index in range(1, 10)])
        self.assertEqual(sheet["columns"]["theirsOnly"], ["@theirs+H"])
        inserted_column = next(
            item for item in sheet["columnAlignment"] if item["field"] == "@theirs+H"
        )
        self.assertEqual(inserted_column["status"], "added_theirs")
        self.assertEqual(inserted_column["anchorAfterBaseColumn"], "G")
        self.assertEqual(sheet["rowCounts"], {"base": 10, "ours": 10, "theirs": 11})

        conflicts = [segment for segment in sheet["segments"] if segment["type"] == "conflict"]
        self.assertEqual(len(conflicts), 1)
        self.assertFalse(conflicts[0]["alignmentRequired"])
        self.assertEqual(len(conflicts[0]["rowSlots"]), 1)
        diffs = conflicts[0]["rowSlots"][0]["cellDiffs"]
        self.assertEqual(
            [cell["field"] for cell in diffs if cell["resolution"] == "conflict"],
            [f"@{index}" for index in range(1, 9)],
        )
        self.assertEqual(
            [cell["field"] for cell in diffs if cell["resolution"] == "ours"],
            ["@9"],
        )
        self.assertIn([0] * 9, merged_rows(sheet))

    def test_headerless_multiple_middle_columns_and_rows_merge_cleanly(self):
        write_raw_book(self.base, [
            [1, 10, 100, 1000],
            [2, 20, 200, 2000],
            [3, 30, 300, 3000],
        ])
        write_raw_book(self.ours, [
            [1, 10, "x1", "y1", 100, 1000],
            [2, 20, "x2", "y2", 200, 2000],
            [3, 30, "x3", "y3", 300, 3000],
        ])
        write_raw_book(self.theirs, [
            [1, 10, 100, 1000],
            [11, 110, 1100, 11000],
            [12, 120, 1200, 12000],
            [2, 20, 200, 2000],
            [3, 30, 300, 3001],
        ])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(sheet["projectionFields"], ["@1", "@2", "@3", "@4"])
        self.assertEqual(sheet["columns"]["oursOnly"], ["@ours+C", "@ours+D"])
        self.assertEqual(
            merged_rows(sheet),
            [
                [1, 10, 100, 1000],
                [11, 110, 1100, 11000],
                [12, 120, 1200, 12000],
                [2, 20, 200, 2000],
                [3, 30, 300, 3001],
            ],
        )

    def test_both_add_different_columns_at_same_anchor_preserves_local_then_remote(self):
        write_raw_book(self.base, [[1, 10], [2, 20]])
        write_raw_book(self.ours, [[1, "local-a", 10], [2, "local-b", 20]])
        write_raw_book(self.theirs, [[1, "remote-a", 10], [2, "remote-b", 20]])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(
            sheet["columns"]["all"],
            ["@1", "@ours+B", "@theirs+B", "@2"],
        )
        added = [
            item["status"] for item in sheet["columnAlignment"]
            if item["status"].startswith("added_")
        ]
        self.assertEqual(added, ["added_ours", "added_theirs"])

    def test_both_add_identical_column_at_same_anchor_collapses_it(self):
        write_raw_book(self.base, [[1, 10], [2, 20]])
        added = [[1, "same-a", 10], [2, "same-b", 20]]
        write_raw_book(self.ours, added)
        write_raw_book(self.theirs, added)

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        column = next(
            item for item in sheet["columnAlignment"]
            if item["status"] == "added_both_same"
        )
        self.assertEqual(column["oursColumn"], "B")
        self.assertEqual(column["theirsColumn"], "B")
        self.assertEqual(len(sheet["columns"]["all"]), 3)

    def test_one_side_adds_column_and_deletes_another(self):
        write_raw_book(self.base, [[1, 10, 100], [2, 20, 200]])
        write_raw_book(self.ours, [[1, 100, "new-a"], [2, 200, "new-b"]])
        write_raw_book(self.theirs, [[1, 10, 100], [2, 20, 200]])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        statuses = {item["field"]: item["status"] for item in sheet["columnAlignment"]}
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(statuses["@2"], "deleted_ours")
        self.assertIn("added_ours", statuses.values())

    def test_sheet_order_is_ignored(self):
        sheets = [("A", [[1]]), ("B", [[2]])]
        write_raw_sheets(self.base, sheets)
        write_raw_sheets(self.ours, list(reversed(sheets)))
        write_raw_sheets(self.theirs, sheets)

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        self.assertEqual(report["summary"]["sheetConflictCount"], 0)
        self.assertEqual([sheet["name"] for sheet in report["sheets"]], ["A", "B"])
        self.assertTrue(all(sheet["renameStatus"] == "unchanged" for sheet in report["sheets"]))

    def test_one_sided_sheet_rename_is_detected_by_content(self):
        write_raw_sheets(self.base, [("Data", [[1], [2]])])
        write_raw_sheets(self.ours, [("Renamed", [[1], [2]])])
        write_raw_sheets(self.theirs, [("Data", [[1], [2]])])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["names"], {
            "base": "Data", "ours": "Renamed", "theirs": "Data"
        })
        self.assertEqual(sheet["renameStatus"], "renamed_ours")
        self.assertEqual(sheet["status"], "clean")

    def test_two_different_sheet_renames_are_a_sheet_conflict(self):
        write_raw_sheets(self.base, [("Data", [[1], [2]])])
        write_raw_sheets(self.ours, [("LocalName", [[1], [2]])])
        write_raw_sheets(self.theirs, [("RemoteName", [[1], [2]])])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        self.assertEqual(sheet["renameStatus"], "rename_conflict")
        self.assertEqual(sheet["status"], "conflict")
        self.assertEqual(report["summary"]["sheetConflictCount"], 1)

    def test_same_name_sheet_added_on_both_sides_uses_remote_without_base(self):
        write_raw_sheets(self.base, [("Keep", [[0]])])
        write_raw_sheets(self.ours, [("Keep", [[0]]), ("New", [[1]])])
        write_raw_sheets(self.theirs, [("Keep", [[0]]), ("New", [[2]])])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = next(item for item in report["sheets"] if item["name"] == "New")
        self.assertEqual(sheet["action"], "added_both")
        self.assertEqual(sheet["resolution"], "theirs")
        self.assertEqual(sheet["policy"], "no_base_remote_wins")

    def test_deleted_sheet_conflicts_with_other_side_modification(self):
        write_raw_sheets(self.base, [("Keep", [[0]]), ("Data", [[1]])])
        write_raw_sheets(self.ours, [("Keep", [[0]])])
        write_raw_sheets(self.theirs, [("Keep", [[0]]), ("Data", [[2]])])

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = next(item for item in report["sheets"] if item["name"] == "Data")
        self.assertEqual(sheet["status"], "conflict")
        self.assertEqual(sheet["sheetConflict"]["kind"], "delete_modify_ours")

    def test_repeated_rows_with_simultaneous_column_changes_keep_positional_identity(self):
        standard = [1, 2, 3, 5, 6, 7, 0, 8, 9]
        base_rows = [standard, standard, standard, [0] * 9, *([standard] * 6), [9] * 9]
        ours_rows = [
            [1, 2, "?", 5, 5, 6, 7, 8, 9],
            *([[1, 2, "?", 4, 5, 6, 7, 8, 9]] * 4),
            *([[1, 2, 3, 4, 5, 6, 7, 8, 9]] * 4),
            [10] * 9,
        ]
        theirs_rows = [
            [*row[:2], "#", *row[3:], None, None]
            for row in base_rows
        ]
        theirs_rows.append([None] * 10 + [10])
        write_raw_book(self.base, base_rows)
        write_raw_book(self.ours, ours_rows)
        write_raw_book(self.theirs, theirs_rows)

        report = analyze_workbooks(self.base, self.ours, self.theirs, self.output)
        sheet = report["sheets"][0]
        columns = [
            (entry["status"], entry.get("baseColumn"), entry.get("oursColumn"), entry.get("theirsColumn"))
            for entry in sheet["columnAlignment"]
        ]
        self.assertEqual(columns, [
            ("common", "A", "A", "A"),
            ("common", "B", "B", "B"),
            ("common", "C", "C", "C"),
            ("added_ours", None, "D", None),
            ("common", "D", "E", "D"),
            ("common", "E", "F", "E"),
            ("common", "F", "G", "F"),
            ("deleted_ours", "G", None, "G"),
            ("common", "H", "H", "H"),
            ("common", "I", "I", "I"),
            ("added_theirs", None, None, "J"),
            ("added_theirs", None, None, "K"),
        ])
        rows = [
            (entry["status"], entry.get("baseRow"), entry.get("oursRow"), entry.get("theirsRow"))
            for entry in sheet["rowAlignment"]
        ]
        self.assertEqual(rows, [
            ("common", 1, 1, 1),
            ("common", 2, 2, 2),
            ("common", 3, 3, 3),
            ("delete_modify_ours", 4, None, 4),
            ("common", 5, 4, 5),
            ("common", 6, 5, 6),
            ("common", 7, 6, 7),
            ("common", 8, 7, 8),
            ("common", 9, 8, 9),
            ("common", 10, 9, 10),
            ("delete_modify_ours", 11, None, 11),
            ("added_ours", None, 10, None),
            ("added_theirs", None, None, 12),
        ])

    def test_cli_writes_report_using_bundled_interpreter_contract(self):
        fields = ["id", "name"]
        rows = [{"id": 1, "name": "a"}]
        write_book(self.base, fields, rows)
        write_book(self.ours, fields, rows)
        write_book(self.theirs, fields, rows)
        script = PACKAGE_ROOT / "xlsx_merge_engine" / "xlsx_git_merge_bridge.py"
        completed = subprocess.run(
            [
                sys.executable, str(script),
                "--base", str(self.base),
                "--ours", str(self.ours),
                "--theirs", str(self.theirs),
                "--output-dir", str(self.output),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=30,
        )
        self.assertEqual(completed.returncode, 0, completed.stdout + completed.stderr)
        result = json.loads(completed.stdout)
        self.assertTrue(result["ok"])
        self.assertTrue(Path(result["report"]).is_file())

    def test_theirs_inserted_empty_row_is_preserved_in_merge(self):
        fields = ["id", "name", "value"]
        rows = [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "c", "value": 30},
            {"id": 3, "name": "d", "value": 40},
        ]
        write_book(self.base, fields, rows)
        write_book(self.ours, fields, rows)
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": None, "name": None, "value": None},
            {"id": 2, "name": "c", "value": 30},
            {"id": 3, "name": "d", "value": 40},
        ])

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(merged_rows(sheet), [["a", 10], [None, None], ["c", 30], ["d", 40]])
        self.assertTrue(any(
            entry.get("status") == "added_theirs" and entry.get("oursRow") is None
            for entry in sheet["rowAlignment"]
        ))

    def test_both_add_empty_rows_at_same_anchor_keeps_union(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": None, "name": None, "value": None},
            {"id": 2, "name": "b", "value": 20},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": None, "name": None, "value": None},
            {"id": None, "name": None, "value": None},
            {"id": 2, "name": "b", "value": 20},
        ])

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(merged_rows(sheet), [["a", 10], [None, None], [None, None], ["b", 20]])

    def test_cleared_data_row_becomes_empty_row_and_merges_cleanly(self):
        fields = ["id", "name", "value"]
        write_book(self.base, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
            {"id": 3, "name": "c", "value": 30},
        ])
        write_book(self.ours, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": None, "name": None, "value": None},
            {"id": 3, "name": "c", "value": 30},
        ])
        write_book(self.theirs, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
            {"id": 3, "name": "c", "value": 30},
        ])

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(merged_rows(sheet), [["a", 10], [None, None], ["c", 30]])

    def test_large_table_layered_alignment_matches_full_dp(self):
        # 合成大表：唯一 key + 修改/插入/删除行，分层路径与全量 DP 输出内容一致
        n = 1200
        base_rows = [
            RowProjection(i, (f"key{i:05d}", "a", i), {"key": f"key{i:05d}"})
            for i in range(n)
        ]
        side_rows = [
            RowProjection(i, (f"key{i:05d}", "a", i), {"key": f"key{i:05d}"})
            for i in range(n)
        ]
        # 修改 30 行、插入 5 行（含同 key 与全新 key）、删除 3 行
        for i in (10, 100, 500, 999, 1100):
            side_rows[i] = RowProjection(
                i, (f"key{i:05d}", "b", i), {"key": f"key{i:05d}"}
            )
        for i in range(25, 55):
            side_rows[i] = RowProjection(
                i, (f"key{i:05d}", "c", i), {"key": f"key{i:05d}"}
            )
        insertions = [
            RowProjection(9000 + k, (f"new{k}", "a", k), {"key": f"new{k}"})
            for k in range(5)
        ]
        side_rows = side_rows[:300] + insertions[:2] + side_rows[300:600] + insertions[2:] + side_rows[600:]
        del side_rows[50], side_rows[700], side_rows[1000]

        def output_sequence(rows_b, rows_s, matches, ins):
            seq = []
            for anchor in range(len(rows_b) + 1):
                for si in ins.get(anchor, ()):
                    seq.append(rows_s[si].projection)
                if anchor < len(rows_b) and anchor in matches:
                    seq.append(rows_s[matches[anchor]].projection)
            return seq

        original_threshold = bridge_mod._FAST_ALIGN_THRESHOLD
        try:
            # 全量路径
            bridge_mod._FAST_ALIGN_THRESHOLD = 10 ** 12
            full_matches, full_ins = _row_side_alignment(base_rows, side_rows)
            full_seq = output_sequence(base_rows, side_rows, full_matches, full_ins)
            # 分层路径
            bridge_mod._FAST_ALIGN_THRESHOLD = original_threshold
            fast_matches, fast_ins = _row_side_alignment(base_rows, side_rows)
            fast_seq = output_sequence(base_rows, side_rows, fast_matches, fast_ins)
        finally:
            bridge_mod._FAST_ALIGN_THRESHOLD = original_threshold

        self.assertEqual(len(full_matches), len(fast_matches))
        self.assertEqual(full_seq, fast_seq)

    def test_large_table_with_repeated_rows_layered_matches_full_dp(self):
        # 重复行 + 修改行混合：分层路径与全量 DP 的合并输出内容一致
        base_rows = [
            RowProjection(i, (f"dup{i % 3}", "a", i), {"key": f"dup{i % 3}"})
            for i in range(1300)
        ]
        side_rows = list(base_rows)
        for i in (1, 5, 400, 900):
            side_rows[i] = RowProjection(i, (f"dup{i % 3}", "b", i), {"key": f"dup{i % 3}"})
        side_rows.insert(500, RowProjection(9000, ("dup1", "c", 1), {"key": "dup1"}))
        del side_rows[1200]

        def output_sequence(rows_b, rows_s, matches, ins):
            seq = []
            for anchor in range(len(rows_b) + 1):
                for si in ins.get(anchor, ()):
                    seq.append(rows_s[si].projection)
                if anchor < len(rows_b) and anchor in matches:
                    seq.append(rows_s[matches[anchor]].projection)
            return seq

        original_threshold = bridge_mod._FAST_ALIGN_THRESHOLD
        try:
            bridge_mod._FAST_ALIGN_THRESHOLD = 10 ** 12
            full_matches, full_ins = _row_side_alignment(base_rows, side_rows)
            full_seq = output_sequence(base_rows, side_rows, full_matches, full_ins)
            bridge_mod._FAST_ALIGN_THRESHOLD = original_threshold
            fast_matches, fast_ins = _row_side_alignment(base_rows, side_rows)
            fast_seq = output_sequence(base_rows, side_rows, fast_matches, fast_ins)
        finally:
            bridge_mod._FAST_ALIGN_THRESHOLD = original_threshold

        self.assertEqual(full_seq, fast_seq)

    def test_column_similarity_stays_within_unit_range(self):
        # 高频重复 token 场景：相似度不得超过 1.0
        short = ("a", "b")
        long = ("a", "a", "a", "b", "b", "b")
        score = bridge_mod._column_similarity(short, long)
        self.assertLessEqual(score, 1.0)
        self.assertGreater(score, 0.0)
        self.assertEqual(bridge_mod._column_similarity(short, short), 1.0)
        self.assertEqual(bridge_mod._column_similarity(short, ("x", "y")), 0.0)

    def test_large_table_layered_alignment_smoke_performance(self):
        n = 5000
        base_rows = [
            RowProjection(i, (f"key{i:05d}", "a", i), {"key": f"key{i:05d}"})
            for i in range(n)
        ]
        side_rows = list(base_rows)
        for i in range(2000, 2020):
            side_rows[i] = RowProjection(
                i, (f"key{i:05d}", "b", i), {"key": f"key{i:05d}"}
            )
        side_rows.append(RowProjection(n, ("tail", "a", n), {"key": "tail"}))

        import time
        start = time.perf_counter()
        matches, insertions = _row_side_alignment(base_rows, side_rows)
        elapsed = time.perf_counter() - start
        self.assertLess(elapsed, 10.0, f"layered alignment too slow: {elapsed:.1f}s")
        self.assertEqual(len(matches), n - 0)  # 全部 base 行匹配（含修改行）

    def test_base_empty_row_deleted_by_ours_is_resolved(self):
        fields = ["id", "name", "value"]
        base_rows = [
            {"id": 1, "name": "a", "value": 10},
            {"id": None, "name": None, "value": None},
            {"id": 2, "name": "b", "value": 20},
        ]
        write_book(self.base, fields, base_rows)
        write_book(self.ours, fields, [
            {"id": 1, "name": "a", "value": 10},
            {"id": 2, "name": "b", "value": 20},
        ])
        write_book(self.theirs, fields, base_rows)

        report = analyze_workbooks(
            self.base, self.ours, self.theirs, self.output, ignore_columns=["id"]
        )
        sheet = report["sheets"][0]
        self.assertEqual(sheet["status"], "clean")
        self.assertEqual(merged_rows(sheet), [["a", 10], ["b", 20]])


if __name__ == "__main__":
    unittest.main()
