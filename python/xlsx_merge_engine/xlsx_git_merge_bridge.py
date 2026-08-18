"""Debug bridge: project three XLSX versions to JSONL and let Git merge rows.

This tool is intentionally independent from the production XLSX resolver.  It
uses ``git merge-file`` only to segment clean row edits from candidate conflict
blocks.  The generated sidecars retain workbook row numbers and complete cell
values so later experiments can map text results back to XLSX safely.
"""

from __future__ import annotations

import argparse
from difflib import SequenceMatcher
import hashlib
import json
import re
import subprocess
import sys
from dataclasses import dataclass
from functools import lru_cache
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MISSING = object()


class BridgeError(RuntimeError):
    """Raised when an XLSX projection or Git text merge cannot be completed."""


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return str(value)


def _line(values: Iterable[Any]) -> str:
    return json.dumps(list(values), ensure_ascii=False, separators=(",", ":"))


def _hash_line(line: str) -> str:
    return hashlib.sha256(line.encode("utf-8")).hexdigest()


def _safe_name(name: str, index: int) -> str:
    slug = re.sub(r"[^0-9A-Za-z._-]+", "_", name).strip("._") or "sheet"
    return f"{index:02d}-{slug[:80]}"


@dataclass(frozen=True)
class SheetLayout:
    name: str
    var_row: int | None
    data_start_row: int
    fields: tuple[str, ...]
    columns: dict[str, int]


@dataclass(frozen=True)
class RowProjection:
    xlsx_row: int
    projection: tuple[Any, ...]
    all_cells: dict[str, Any]
    empty: bool = False

    @property
    def text(self) -> str:
        return _line(self.projection)


@dataclass(frozen=True)
class PairAlignment:
    matches: dict[int, int]
    similarities: dict[int, float]
    side_only: tuple[int, ...]
    base_only: tuple[int, ...]


def _last_semantic_column(ws) -> int:
    """Ignore styled trailing placeholders but retain every column inside the used range."""
    last = 1
    for cell in ws._cells.values():
        if (
            cell.value not in (None, "")
            or getattr(cell, "comment", None) is not None
            or getattr(cell, "hyperlink", None) is not None
        ):
            last = max(last, cell.column)
    for merged_range in ws.merged_cells.ranges:
        last = max(last, merged_range.max_col)
    return last


def _layout(ws) -> SheetLayout:
    var_row = None
    for row in range(1, min(ws.max_row or 1, 40) + 1):
        if str(ws.cell(row, 1).value or "").strip() == "##var":
            var_row = row
            break

    columns: dict[str, int] = {}
    if var_row is None:
        for column in range(1, _last_semantic_column(ws) + 1):
            columns[f"@{column}"] = column
        return SheetLayout(ws.title, None, 1, tuple(columns), columns)

    columns["@1"] = 1
    occurrences: dict[str, int] = {}
    for column in range(2, _last_semantic_column(ws) + 1):
        raw_name = ws.cell(var_row, column).value
        if raw_name is None or str(raw_name).strip() == "":
            columns[f"@{column}"] = column
            continue
        base_name = str(raw_name).strip()
        occurrences[base_name] = occurrences.get(base_name, 0) + 1
        suffix = occurrences[base_name]
        name = base_name if suffix == 1 else f"{base_name}#{suffix}"
        columns[name] = column

    row = var_row
    while row <= (ws.max_row or var_row):
        marker = ws.cell(row, 1).value
        if not isinstance(marker, str) or not marker.startswith("##"):
            break
        row += 1
    return SheetLayout(ws.title, var_row, row, tuple(columns), columns)


def _active_rows(ws, layout: SheetLayout) -> list[int]:
    return [
        row
        for row in range(layout.data_start_row, (ws.max_row or 0) + 1)
        if any(ws.cell(row, column).value is not None for column in layout.columns.values())
    ]


@lru_cache(maxsize=65536)
def _token_string(value: Any) -> str:
    return _line((_json_value(value),))


def _column_tokens(ws, rows: Iterable[int], column: int) -> tuple[str, ...]:
    return tuple(_token_string(ws.cell(row, column).value) for row in rows)


def _sheet_column_tokens(
    ws, active_rows: list[int], columns: list[int]
) -> list[tuple[str, ...]]:
    """一次 iter_rows 读取整表，按列生成 token 序列（大表避免逐 cell 访问）。"""
    if not active_rows or not columns:
        return [() for _ in columns]
    row_set = set(active_rows)
    first = min(active_rows)
    last = max(active_rows)
    buckets: list[list[str]] = [[] for _ in columns]
    for offset, values in enumerate(
        ws.iter_rows(min_row=first, max_row=last, values_only=True), first
    ):
        if offset not in row_set:
            continue
        for bucket, column in zip(buckets, columns):
            bucket.append(
                _token_string(values[column - 1] if values is not None else None)
            )
    return [tuple(bucket) for bucket in buckets]


def _column_similarity(left: tuple[str, ...], right: tuple[str, ...]) -> float:
    if not left and not right:
        return 1.0
    if not left or not right:
        return 0.0
    # 频次配对（线性）：SequenceMatcher 在长序列且大部分不同的场景 O(n*m) 退化
    counts: dict[str, int] = {}
    for token in left:
        counts[token] = counts.get(token, 0) + 1
    matched = 0
    for token in right:
        if counts.get(token, 0):
            counts[token] -= 1
            matched += 1
    # 上限截断：公共 token 数不超过 min 长度，保证相似度 ∈ [0, 1]
    return min(matched, len(left), len(right)) / min(len(left), len(right))


def _align_column_sequences(
    base_columns: list[tuple[str, ...]],
    side_columns: list[tuple[str, ...]],
    *,
    minimum_similarity: float = 0.0,
) -> PairAlignment:
    """Globally align ordered columns while allowing inserts and deletes."""
    base_count = len(base_columns)
    side_count = len(side_columns)
    gap = -0.40
    negative = float("-inf")
    scores = [[negative] * (side_count + 1) for _ in range(base_count + 1)]
    moves: list[list[str | None]] = [
        [None] * (side_count + 1) for _ in range(base_count + 1)
    ]
    scores[0][0] = 0.0
    for base_index in range(1, base_count + 1):
        scores[base_index][0] = base_index * gap
        moves[base_index][0] = "delete"
    for side_index in range(1, side_count + 1):
        scores[0][side_index] = side_index * gap
        moves[0][side_index] = "insert"

    similarities = [
        [
            _column_similarity(base_columns[base_index], side_columns[side_index])
            for side_index in range(side_count)
        ]
        for base_index in range(base_count)
    ]
    priority = {"match": 2, "delete": 1, "insert": 0}
    for base_index in range(1, base_count + 1):
        for side_index in range(1, side_count + 1):
            candidates = [
                (scores[base_index - 1][side_index] + gap, "delete"),
                (scores[base_index][side_index - 1] + gap, "insert"),
            ]
            similarity = similarities[base_index - 1][side_index - 1]
            if similarity >= minimum_similarity:
                candidates.append(
                    (scores[base_index - 1][side_index - 1] + similarity, "match")
                )
            score, move = max(candidates, key=lambda item: (item[0], priority[item[1]]))
            scores[base_index][side_index] = score
            moves[base_index][side_index] = move

    matches: dict[int, int] = {}
    match_similarities: dict[int, float] = {}
    base_only: list[int] = []
    side_only: list[int] = []
    base_index = base_count
    side_index = side_count
    while base_index or side_index:
        move = moves[base_index][side_index]
        if move == "match":
            left = base_index
            right = side_index
            matches[left] = right
            match_similarities[left] = similarities[left - 1][right - 1]
            base_index -= 1
            side_index -= 1
        elif move == "delete":
            base_only.append(base_index)
            base_index -= 1
        elif move == "insert":
            side_only.append(side_index)
            side_index -= 1
        else:
            raise BridgeError("Column sequence alignment backtracking failed.")

    base_only_set = set(base_only)
    side_only_set = set(side_only)
    anchors = [(0, 0), *sorted(matches.items()), (base_count + 1, side_count + 1)]
    for (left_base, left_side), (right_base, right_side) in zip(anchors, anchors[1:]):
        base_gap = [
            index for index in range(left_base + 1, right_base)
            if index in base_only_set
        ]
        side_gap = [
            index for index in range(left_side + 1, right_side)
            if index in side_only_set
        ]
        if not base_gap or len(base_gap) != len(side_gap):
            continue
        for base_position, side_position in zip(base_gap, side_gap):
            matches[base_position] = side_position
            match_similarities[base_position] = similarities[base_position - 1][side_position - 1]
            base_only_set.remove(base_position)
            side_only_set.remove(side_position)

    return PairAlignment(
        matches=matches,
        similarities=match_similarities,
        side_only=tuple(sorted(side_only_set)),
        base_only=tuple(sorted(base_only_set)),
    )


def _insertion_anchor(column: int, matches: dict[int, int]) -> int:
    return max((base for base, side in matches.items() if side < column), default=0)


def _logical_insert_name(side: str, physical_column: int) -> str:
    return f"@{side}+{get_column_letter(physical_column)}"


def _is_positional_field(field: str) -> bool:
    return re.fullmatch(r"@\d+", field) is not None


def _align_content_layouts(
    worksheets: dict[str, Any], layouts: dict[str, SheetLayout]
) -> tuple[dict[str, SheetLayout], list[dict[str, Any]]]:
    entries = {
        side: [(field, layouts[side].columns[field]) for field in layouts[side].fields]
        for side in ("base", "ours", "theirs")
    }
    active_rows = {
        side: _active_rows(worksheets[side], layouts[side])
        for side in ("base", "ours", "theirs")
    }
    vectors = {
        side: _sheet_column_tokens(
            worksheets[side],
            active_rows[side],
            [column for _, column in entries[side]],
        )
        for side in ("base", "ours", "theirs")
    }
    pair = {
        side: _align_column_sequences(vectors["base"], vectors[side])
        for side in ("ours", "theirs")
    }
    aligned_columns: dict[str, dict[str, int]] = {
        "base": {field: column for field, column in entries["base"]},
        "ours": {},
        "theirs": {},
    }
    column_report: list[dict[str, Any]] = []
    for base_position, (field, base_physical) in enumerate(entries["base"], 1):
        ours_position = pair["ours"].matches.get(base_position)
        theirs_position = pair["theirs"].matches.get(base_position)
        ours_physical = entries["ours"][ours_position - 1][1] if ours_position else None
        theirs_physical = entries["theirs"][theirs_position - 1][1] if theirs_position else None
        ours_label = entries["ours"][ours_position - 1][0] if ours_position else None
        theirs_label = entries["theirs"][theirs_position - 1][0] if theirs_position else None
        if ours_physical is not None:
            aligned_columns["ours"][field] = ours_physical
        if theirs_physical is not None:
            aligned_columns["theirs"][field] = theirs_physical
        if ours_physical is None and theirs_physical is None:
            status = "deleted_both"
        elif ours_physical is None:
            status = "deleted_ours"
        elif theirs_physical is None:
            status = "deleted_theirs"
        else:
            status = "common"
        column_report.append({
            "field": field,
            "status": status,
            "labels": {"base": field, "ours": ours_label, "theirs": theirs_label},
            "baseColumn": get_column_letter(base_physical),
            "oursColumn": get_column_letter(ours_physical) if ours_physical else None,
            "theirsColumn": get_column_letter(theirs_physical) if theirs_physical else None,
            "oursSimilarity": pair["ours"].similarities.get(base_position),
            "theirsSimilarity": pair["theirs"].similarities.get(base_position),
        })

    unmatched = {
        side: [
            (column, _insertion_anchor(column, pair[side].matches))
            for column in pair[side].side_only
        ]
        for side in ("ours", "theirs")
    }
    anchors = sorted({anchor for values in unmatched.values() for _, anchor in values})
    inserted_fields: dict[int, list[str]] = {}
    used_fields = set(aligned_columns["base"])

    def inserted_field(side: str, position: int) -> str:
        label, physical = entries[side][position - 1]
        if not _is_positional_field(label) and label not in used_fields:
            used_fields.add(label)
            return label
        candidate = _logical_insert_name(side, physical)
        suffix = 2
        while candidate in used_fields:
            candidate = f"{_logical_insert_name(side, physical)}#{suffix}"
            suffix += 1
        used_fields.add(candidate)
        return candidate

    for anchor in anchors:
        ours_positions = [column for column, value in unmatched["ours"] if value == anchor]
        theirs_positions = [column for column, value in unmatched["theirs"] if value == anchor]
        matched_theirs: set[int] = set()
        for ours_position in ours_positions:
            theirs_position = next((
                candidate for candidate in theirs_positions
                if candidate not in matched_theirs
                and vectors["ours"][ours_position - 1] == vectors["theirs"][candidate - 1]
            ), None)
            if theirs_position is None:
                label, physical = entries["ours"][ours_position - 1]
                field = inserted_field("ours", ours_position)
                inserted_fields.setdefault(anchor, []).append(field)
                aligned_columns["ours"][field] = physical
                column_report.append({
                    "field": field,
                    "status": "added_ours",
                    "labels": {"base": None, "ours": label},
                    "baseColumn": None,
                    "oursColumn": get_column_letter(physical),
                    "theirsColumn": None,
                    "anchorAfterBaseColumn": get_column_letter(entries["base"][anchor - 1][1]) if anchor else None,
                })
                continue
            ours_label, ours_physical = entries["ours"][ours_position - 1]
            theirs_label, theirs_physical = entries["theirs"][theirs_position - 1]
            if ours_label == theirs_label and not _is_positional_field(ours_label) and ours_label not in used_fields:
                field = ours_label
                used_fields.add(field)
            else:
                field = f"@both+{anchor + 1}.{len(matched_theirs)}"
                used_fields.add(field)
            inserted_fields.setdefault(anchor, []).append(field)
            aligned_columns["ours"][field] = ours_physical
            aligned_columns["theirs"][field] = theirs_physical
            column_report.append({
                "field": field,
                "status": "added_both_same",
                "labels": {"base": None, "ours": ours_label, "theirs": theirs_label},
                "baseColumn": None,
                "oursColumn": get_column_letter(ours_physical),
                "theirsColumn": get_column_letter(theirs_physical),
                "oursTheirsSimilarity": 1.0,
                "anchorAfterBaseColumn": get_column_letter(entries["base"][anchor - 1][1]) if anchor else None,
            })
            matched_theirs.add(theirs_position)
        for theirs_position in theirs_positions:
            if theirs_position in matched_theirs:
                continue
            label, physical = entries["theirs"][theirs_position - 1]
            field = inserted_field("theirs", theirs_position)
            inserted_fields.setdefault(anchor, []).append(field)
            aligned_columns["theirs"][field] = physical
            column_report.append({
                "field": field,
                "status": "added_theirs",
                "labels": {"base": None, "theirs": label},
                "baseColumn": None,
                "oursColumn": None,
                "theirsColumn": get_column_letter(physical),
                "anchorAfterBaseColumn": get_column_letter(entries["base"][anchor - 1][1]) if anchor else None,
            })

    logical_order = list(inserted_fields.get(0, ()))
    for base_position, (field, _) in enumerate(entries["base"], 1):
        logical_order.append(field)
        logical_order.extend(inserted_fields.get(base_position, ()))
    aligned_layouts = {}
    for side in ("base", "ours", "theirs"):
        columns = aligned_columns[side]
        fields = tuple(field for field in logical_order if field in columns)
        source = layouts[side]
        aligned_layouts[side] = SheetLayout(
            source.name, source.var_row, source.data_start_row, fields, columns
        )
    report_by_field = {entry["field"]: entry for entry in column_report}
    return aligned_layouts, [report_by_field[field] for field in logical_order]


def _align_layouts(
    worksheets: dict[str, Any], layouts: dict[str, SheetLayout]
) -> tuple[dict[str, SheetLayout], list[dict[str, Any]], str]:
    aligned, report = _align_content_layouts(worksheets, layouts)
    return aligned, report, "content_anchor_sequence"


def _rows(ws, layout: SheetLayout, projection_fields: tuple[str, ...]) -> list[RowProjection]:
    result: list[RowProjection] = []
    columns = [(field, column) for field, column in layout.columns.items()]
    for offset, values in enumerate(
        ws.iter_rows(
            min_row=layout.data_start_row,
            max_row=ws.max_row or layout.data_start_row,
            values_only=True,
        ),
        layout.data_start_row,
    ):
        all_cells = {
            field: _json_value(values[column - 1] if values is not None else None)
            for field, column in columns
        }
        projection = tuple(all_cells.get(field) for field in projection_fields)
        if not any(value is not None for value in all_cells.values()):
            # 空行保留：投影全 None，empty 标记供对齐判断
            result.append(RowProjection(offset, projection, all_cells, empty=True))
            continue
        result.append(RowProjection(offset, projection, all_cells))
    return result


def _paired_projected_rows(
    base_rows: list[RowProjection], side_rows: list[RowProjection]
) -> tuple[list[tuple[RowProjection, RowProjection]], list[RowProjection]]:
    """Pair preserved/replaced rows and return rows inserted only on one side."""
    matches, insertions = _row_side_alignment(base_rows, side_rows)
    pairs = [
        (base_rows[base_index], side_rows[side_index])
        for base_index, side_index in sorted(matches.items())
    ]
    inserted = [
        side_rows[side_index]
        for anchor in sorted(insertions)
        for side_index in insertions[anchor]
    ]
    return pairs, inserted


def _row_similarity(left: RowProjection, right: RowProjection) -> float:
    if not left.projection and not right.projection:
        return 1.0
    size = max(len(left.projection), len(right.projection))
    if not size:
        return 0.0
    limit = min(len(left.projection), len(right.projection))
    equal = 0
    for index in range(limit):
        if left.projection[index] == right.projection[index]:
            equal += 1
    return equal / size


_FAST_ALIGN_THRESHOLD = 1_000_000  # base*side 超过该值走分层快速对齐
_SIMILAR_MATCH_MIN = 0.3  # match 分 2s-1 > gap(-0.4) 的临界相似度


def _similar_anchors_in_gap(
    base_rows: list[RowProjection],
    side_rows: list[RowProjection],
    base_gap: list[int],
    side_gap: list[int],
    matches: dict[int, int],
    used_base: set[int],
    used_side: set[int],
) -> None:
    """大区间退化场景：空行同序配对 + 候选剪枝相似锚点，其余按删+插。"""
    b_empty = [i for i in base_gap if base_rows[i].empty]
    s_empty = [j for j in side_gap if side_rows[j].empty]
    for i, j in zip(b_empty, s_empty):
        matches[i] = j
        used_base.add(i)
        used_side.add(j)
    index: dict[tuple[int, Any], list[int]] = {}
    for j in side_gap:
        if j in used_side:
            continue
        for p, v in enumerate(side_rows[j].projection):
            if v is None:
                continue
            index.setdefault((p, v), []).append(j)
    side_gap_set = set(side_gap)
    last_side = max(
        (j for _, j in matches.items() if j in used_side and j in side_gap_set),
        default=-1,
    )
    for i in base_gap:
        if i in used_base:
            continue
        candidates: set[int] = set()
        for p, v in enumerate(base_rows[i].projection):
            if v is None:
                continue
            candidates.update(index.get((p, v), ()))
        best = None
        best_score = _SIMILAR_MATCH_MIN
        for j in candidates:
            if j in used_side or j <= last_side:
                continue
            score = _row_similarity(base_rows[i], side_rows[j])
            if score >= best_score:
                best_score = score
                best = j
        if best is not None:
            matches[i] = best
            used_base.add(i)
            used_side.add(best)
            last_side = best


def _fast_row_side_alignment(
    base_rows: list[RowProjection], side_rows: list[RowProjection]
) -> tuple[dict[int, int], dict[int, list[int]]]:
    """大数据量分层对齐：hash 锚点切区间 → 区间内局部 DP / 候选剪枝。

    与全量 DP 输出内容等价（配对与删+插的写回结果一致），仅行状态标注可能有差异。
    """
    base_count = len(base_rows)
    side_count = len(side_rows)
    matches: dict[int, int] = {}
    used_base: set[int] = set()
    used_side: set[int] = set()

    # 阶段 1：hash 锚点（投影相同的行直接配对）
    base_hashes = [_hash_line(row.text) for row in base_rows]
    side_hashes = [_hash_line(row.text) for row in side_rows]
    matcher = SequenceMatcher(None, base_hashes, side_hashes, autojunk=False)
    for op, b1, b2, s1, s2 in matcher.get_opcodes():
        if op != "equal":
            continue
        for offset in range(b2 - b1):
            matches[b1 + offset] = s1 + offset
            used_base.add(b1 + offset)
            used_side.add(s1 + offset)

    # 阶段 2：按锚点切区间处理未匹配行
    anchors = [(-1, -1), *sorted(matches.items()), (base_count, side_count)]
    for (left_base, left_side), (right_base, right_side) in zip(anchors, anchors[1:]):
        base_gap = [i for i in range(left_base + 1, right_base) if i not in used_base]
        side_gap = [i for i in range(left_side + 1, right_side) if i not in used_side]
        if not base_gap and not side_gap:
            continue
        if len(base_gap) * len(side_gap) <= _FAST_ALIGN_THRESHOLD:
            # 区间小：局部 DP，与全量语义一致（含空行/低相似同序行）
            local_matches, _ = _row_side_alignment(
                [base_rows[i] for i in base_gap], [side_rows[j] for j in side_gap]
            )
            for local_base, local_side in local_matches.items():
                matches[base_gap[local_base]] = side_gap[local_side]
                used_base.add(base_gap[local_base])
                used_side.add(side_gap[local_side])
            continue
        # 区间大（大量行被改）：候选剪枝近似处理
        _similar_anchors_in_gap(
            base_rows, side_rows, base_gap, side_gap, matches, used_base, used_side
        )

    # 阶段 3：插入归属（与全量路径一致）
    insertion_indexes = []
    for inserted_index in range(side_count):
        if inserted_index in used_side:
            continue
        anchor = max(
            (matched_base + 1 for matched_base, matched_side in matches.items() if matched_side < inserted_index),
            default=0,
        )
        insertion_indexes.append((anchor, inserted_index))
    insertions: dict[int, list[int]] = {}
    last_matched_base = max(matches, default=-1)
    last_matched_side = max(matches.values(), default=-1)
    for anchor, inserted_index in insertion_indexes:
        if inserted_index > last_matched_side and anchor > last_matched_base:
            anchor = base_count
        insertions.setdefault(anchor, []).append(inserted_index)
    return matches, insertions


def _row_side_alignment(
    base_rows: list[RowProjection], side_rows: list[RowProjection]
) -> tuple[dict[int, int], dict[int, list[int]]]:
    if len(base_rows) * len(side_rows) > _FAST_ALIGN_THRESHOLD:
        return _fast_row_side_alignment(base_rows, side_rows)
    base_count = len(base_rows)
    side_count = len(side_rows)
    gap = -0.40
    negative = float("-inf")
    scores = [[negative] * (side_count + 1) for _ in range(base_count + 1)]
    moves: list[list[str | None]] = [[None] * (side_count + 1) for _ in range(base_count + 1)]
    scores[0][0] = 0.0
    for base_index in range(1, base_count + 1):
        scores[base_index][0] = base_index * gap
        moves[base_index][0] = "delete"
    for side_index in range(1, side_count + 1):
        scores[0][side_index] = side_index * gap
        moves[0][side_index] = "insert"

    similarities = [
        [_row_similarity(base_row, side_row) for side_row in side_rows]
        for base_row in base_rows
    ]
    priority = {"match": 2, "delete": 1, "insert": 0}
    for base_index in range(1, base_count + 1):
        for side_index in range(1, side_count + 1):
            candidates = [
                (scores[base_index - 1][side_index] + gap, "delete"),
                (scores[base_index][side_index - 1] + gap, "insert"),
                (
                    scores[base_index - 1][side_index - 1]
                    + (2.0 * similarities[base_index - 1][side_index - 1] - 1.0),
                    "match",
                ),
            ]
            score, move = max(candidates, key=lambda item: (item[0], priority[item[1]]))
            scores[base_index][side_index] = score
            moves[base_index][side_index] = move

    matches: dict[int, int] = {}
    base_index = base_count
    side_index = side_count
    while base_index or side_index:
        move = moves[base_index][side_index]
        if move == "match":
            matches[base_index - 1] = side_index - 1
            base_index -= 1
            side_index -= 1
        elif move == "delete":
            base_index -= 1
        elif move == "insert":
            side_index -= 1
        else:
            raise BridgeError("Row sequence alignment backtracking failed.")

    used_base = set(matches)
    used_side = set(matches.values())
    anchors = [(-1, -1), *sorted(matches.items()), (base_count, side_count)]
    for (left_base, left_side), (right_base, right_side) in zip(anchors, anchors[1:]):
        base_gap = [index for index in range(left_base + 1, right_base) if index not in used_base]
        side_gap = [index for index in range(left_side + 1, right_side) if index not in used_side]
        if not base_gap or len(base_gap) != len(side_gap):
            continue
        if any(
            base_rows[left].empty != side_rows[right].empty
            for left, right in zip(base_gap, side_gap)
        ):
            continue
        pair_similarities = [similarities[left][right] for left, right in zip(base_gap, side_gap)]
        if base_count != side_count and not all(similarity > 0.0 for similarity in pair_similarities):
            continue
        for left, right in zip(base_gap, side_gap):
            matches[left] = right
            used_base.add(left)
            used_side.add(right)

    insertion_indexes = []
    for inserted_index in range(side_count):
        if inserted_index in used_side:
            continue
        anchor = max(
            (matched_base + 1 for matched_base, matched_side in matches.items() if matched_side < inserted_index),
            default=0,
        )
        insertion_indexes.append((anchor, inserted_index))

    insertions: dict[int, list[int]] = {}
    last_matched_base = max(matches, default=-1)
    last_matched_side = max(matches.values(), default=-1)
    for anchor, inserted_index in insertion_indexes:
        if inserted_index > last_matched_side and anchor > last_matched_base:
            anchor = base_count
        insertions.setdefault(anchor, []).append(inserted_index)
    return matches, insertions


def _base_row_changed(base: RowProjection, side: RowProjection) -> bool:
    common = set(base.all_cells) & set(side.all_cells)
    return any(base.all_cells[field] != side.all_cells[field] for field in common)


def _row_alignment_report(rows: dict[str, list[RowProjection]]) -> list[dict[str, Any]]:
    mappings = {
        side: _row_side_alignment(rows["base"], rows[side])
        for side in ("ours", "theirs")
    }
    result: list[dict[str, Any]] = []
    insertion_fields: dict[int, list[dict[str, Any]]] = {}
    anchors = sorted(set(mappings["ours"][1]) | set(mappings["theirs"][1]))
    for anchor in anchors:
        ours_indexes = list(mappings["ours"][1].get(anchor, ()))
        theirs_indexes = list(mappings["theirs"][1].get(anchor, ()))
        matched_theirs: set[int] = set()
        for ours_index in ours_indexes:
            ours_row = rows["ours"][ours_index]
            theirs_index = next((
                candidate for candidate in theirs_indexes
                if candidate not in matched_theirs
                and ours_row.projection == rows["theirs"][candidate].projection
            ), None)
            if theirs_index is None:
                insertion_fields.setdefault(anchor, []).append({
                    "id": f"row:ours:{ours_row.xlsx_row}",
                    "status": "added_ours",
                    "baseRow": None,
                    "oursRow": ours_row.xlsx_row,
                    "theirsRow": None,
                    "anchorAfterBaseRow": rows["base"][anchor - 1].xlsx_row if anchor else None,
                })
                continue
            matched_theirs.add(theirs_index)
            theirs_row = rows["theirs"][theirs_index]
            insertion_fields.setdefault(anchor, []).append({
                "id": f"row:both:{ours_row.xlsx_row}:{theirs_row.xlsx_row}",
                "status": "added_both_same",
                "baseRow": None,
                "oursRow": ours_row.xlsx_row,
                "theirsRow": theirs_row.xlsx_row,
                "anchorAfterBaseRow": rows["base"][anchor - 1].xlsx_row if anchor else None,
            })
        for theirs_index in theirs_indexes:
            if theirs_index in matched_theirs:
                continue
            theirs_row = rows["theirs"][theirs_index]
            insertion_fields.setdefault(anchor, []).append({
                "id": f"row:theirs:{theirs_row.xlsx_row}",
                "status": "added_theirs",
                "baseRow": None,
                "oursRow": None,
                "theirsRow": theirs_row.xlsx_row,
                "anchorAfterBaseRow": rows["base"][anchor - 1].xlsx_row if anchor else None,
            })

    result.extend(insertion_fields.get(0, ()))
    for base_index, base_row in enumerate(rows["base"]):
        ours_index = mappings["ours"][0].get(base_index)
        theirs_index = mappings["theirs"][0].get(base_index)
        ours_row = rows["ours"][ours_index] if ours_index is not None else None
        theirs_row = rows["theirs"][theirs_index] if theirs_index is not None else None
        if ours_row is None and theirs_row is None:
            status = "deleted_both"
        elif ours_row is None:
            status = "delete_modify_ours" if _base_row_changed(base_row, theirs_row) else "deleted_ours"
        elif theirs_row is None:
            status = "delete_modify_theirs" if _base_row_changed(base_row, ours_row) else "deleted_theirs"
        else:
            status = "common"
        result.append({
            "id": f"row:base:{base_row.xlsx_row}",
            "status": status,
            "baseRow": base_row.xlsx_row,
            "oursRow": ours_row.xlsx_row if ours_row else None,
            "theirsRow": theirs_row.xlsx_row if theirs_row else None,
        })
        result.extend(insertion_fields.get(base_index + 1, ()))
    return result


def _deleted_column_content_changes(
    field: str,
    base_rows: list[RowProjection],
    side_rows: list[RowProjection],
    side: str,
) -> list[dict[str, Any]]:
    pairs, inserted = _paired_projected_rows(base_rows, side_rows)
    changes: list[dict[str, Any]] = []
    for base_row, side_row in pairs:
        base_value = base_row.all_cells.get(field, MISSING)
        side_value = side_row.all_cells.get(field, MISSING)
        if base_value == side_value:
            continue
        changes.append({
            "kind": f"modified_{side}",
            "baseXlsxRow": base_row.xlsx_row,
            f"{side}XlsxRow": side_row.xlsx_row,
            "base": _display(base_value),
            side: _display(side_value),
        })
    for side_row in inserted:
        side_value = side_row.all_cells.get(field, MISSING)
        if side_value is MISSING or side_value is None:
            continue
        changes.append({
            "kind": f"inserted_{side}",
            "baseXlsxRow": None,
            f"{side}XlsxRow": side_row.xlsx_row,
            "base": None,
            side: _display(side_value),
        })
    return changes


def _annotate_column_conflicts(
    column_alignment: list[dict[str, Any]],
    rows: dict[str, list[RowProjection]],
) -> int:
    conflict_count = 0
    for column in column_alignment:
        status = column["status"]
        if status == "deleted_theirs":
            survivor = "ours"
            deleted = "theirs"
        elif status == "deleted_ours":
            survivor = "theirs"
            deleted = "ours"
        else:
            continue
        changes = _deleted_column_content_changes(
            column["field"], rows["base"], rows[survivor], survivor
        )
        if not changes:
            continue
        column["status"] = f"delete_modify_{deleted}"
        column["conflict"] = True
        column["contentChanges"] = changes
        conflict_count += 1
    return conflict_count


def _write_text_lf(path: Path, content: str) -> None:
    """3.9-compatible LF write (Path.write_text gained `newline` only in 3.10)."""
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write(content)


def _write_projection(path: Path, rows: list[RowProjection]) -> None:
    content = "\n".join(row.text for row in rows)
    if content:
        content += "\n"
    _write_text_lf(path, content)


def _write_sidecar(path: Path, layout: SheetLayout, rows: list[RowProjection]) -> None:
    payload = {
        "sheet": layout.name,
        "varRow": layout.var_row,
        "dataStartRow": layout.data_start_row,
        "fields": list(layout.fields),
        "rows": [
            {
                "textLine": index,
                "xlsxRow": row.xlsx_row,
                "hash": _hash_line(row.text),
                "projection": list(row.projection),
                "cells": row.all_cells,
            }
            for index, row in enumerate(rows, 1)
        ],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _git_merge_file(ours: Path, base: Path, theirs: Path) -> tuple[str, bool, str]:
    result = subprocess.run(
        [
            "git", "merge-file", "-p", "--diff3",
            "-L", "OURS", "-L", "BASE", "-L", "THEIRS",
            str(ours), str(base), str(theirs),
        ],
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=60,
    )
    output = result.stdout.replace("\r\n", "\n")
    has_conflicts = "<<<<<<< OURS" in output
    if result.returncode != 0 and not has_conflicts:
        raise BridgeError(f"git merge-file failed ({result.returncode}): {result.stderr.strip()}")
    return output, not has_conflicts, result.stderr.strip()


def _decode_rows(lines: list[str]) -> list[list[Any]]:
    rows = []
    for line in lines:
        if not line.strip():
            continue
        try:
            value = json.loads(line)
        except json.JSONDecodeError as exc:
            raise BridgeError(f"Merged projection contains invalid JSONL: {line!r}") from exc
        if not isinstance(value, list):
            raise BridgeError(f"Merged projection line is not a JSON array: {line!r}")
        rows.append(value)
    return rows


def _display(value: Any) -> Any:
    return None if value is MISSING else value


def _row_kind(base: Any, ours: Any, theirs: Any) -> str:
    if base is MISSING:
        if ours is MISSING:
            return "inserted_theirs"
        if theirs is MISSING:
            return "inserted_ours"
        return "inserted_both_same" if ours == theirs else "inserted_both_conflict"
    if ours is MISSING and theirs is MISSING:
        return "deleted_both"
    if ours is MISSING:
        return "deleted_ours" if theirs == base else "delete_modify_ours"
    if theirs is MISSING:
        return "deleted_theirs" if ours == base else "delete_modify_theirs"
    if ours == theirs:
        return "same"
    if ours == base:
        return "modified_theirs"
    if theirs == base:
        return "modified_ours"
    return "modified_both"


def _cell_diffs(
    fields: tuple[str, ...],
    base_rows: list[list[Any]],
    ours_rows: list[list[Any]],
    theirs_rows: list[list[Any]],
) -> tuple[list[dict[str, Any]], bool]:
    alignment_required = len({len(base_rows), len(ours_rows), len(theirs_rows)}) > 1
    slots: list[dict[str, Any]] = []
    for index in range(max(len(base_rows), len(ours_rows), len(theirs_rows), 0)):
        base = base_rows[index] if index < len(base_rows) else MISSING
        ours = ours_rows[index] if index < len(ours_rows) else MISSING
        theirs = theirs_rows[index] if index < len(theirs_rows) else MISSING
        differences = []
        for column, field in enumerate(fields):
            base_value = base[column] if base is not MISSING and column < len(base) else MISSING
            ours_value = ours[column] if ours is not MISSING and column < len(ours) else MISSING
            theirs_value = theirs[column] if theirs is not MISSING and column < len(theirs) else MISSING
            if ours_value == theirs_value == base_value:
                continue
            if base_value is MISSING and theirs_value is not MISSING:
                resolution = "theirs"
            elif base_value is MISSING and ours_value is not MISSING:
                resolution = "ours"
            elif ours_value == theirs_value:
                resolution = "same"
            elif ours_value == base_value:
                resolution = "theirs"
            elif theirs_value == base_value:
                resolution = "ours"
            else:
                resolution = "conflict"
            differences.append({
                "field": field,
                "base": _display(base_value),
                "ours": _display(ours_value),
                "theirs": _display(theirs_value),
                "basePresent": base_value is not MISSING,
                "oursPresent": ours_value is not MISSING,
                "theirsPresent": theirs_value is not MISSING,
                "resolution": resolution,
            })
        slots.append({
            "slot": index + 1,
            "kind": _row_kind(base, ours, theirs),
            "base": _display(base),
            "ours": _display(ours),
            "theirs": _display(theirs),
            "cellDiffs": differences,
        })
    return slots, alignment_required


def _merge_inserted_rows(
    ours_rows: list[list[Any]], theirs_rows: list[list[Any]]
) -> list[list[Any]]:
    """Union same-anchor inserts, collapsing cross-side duplicates, local first."""
    resolved = list(ours_rows)
    local_counts: dict[str, int] = {}
    for row in ours_rows:
        key = _line(row)
        local_counts[key] = local_counts.get(key, 0) + 1
    for row in theirs_rows:
        key = _line(row)
        if local_counts.get(key, 0):
            local_counts[key] -= 1
        else:
            resolved.append(row)
    return resolved


def _row_edit_plan(
    base_rows: list[list[Any]], side_rows: list[list[Any]]
) -> tuple[list[Any], dict[int, list[list[Any]]]]:
    states: list[Any] = [MISSING] * len(base_rows)
    insertions: dict[int, list[list[Any]]] = {}
    matcher = SequenceMatcher(
        None,
        [_line(row) for row in base_rows],
        [_line(row) for row in side_rows],
        autojunk=False,
    )
    for operation, base_start, base_end, side_start, side_end in matcher.get_opcodes():
        if operation == "equal":
            for offset, row in enumerate(side_rows[side_start:side_end]):
                states[base_start + offset] = row
        elif operation == "replace":
            base_count = base_end - base_start
            side_block = side_rows[side_start:side_end]
            paired_count = min(base_count, len(side_block))
            for offset, row in enumerate(side_block[:paired_count]):
                states[base_start + offset] = row
            if len(side_block) > paired_count:
                insertions.setdefault(base_start + paired_count, []).extend(side_block[paired_count:])
        elif operation == "insert":
            insertions.setdefault(base_start, []).extend(side_rows[side_start:side_end])
        elif operation != "delete":
            raise BridgeError(f"Unsupported conflict-block row operation: {operation}")
    return states, insertions


def _merge_base_row(
    fields: tuple[str, ...], base: list[Any], ours: Any, theirs: Any
) -> tuple[str, Any]:
    if ours is MISSING and theirs is MISSING:
        return "resolved", MISSING
    if ours is MISSING:
        if theirs == base:
            return "resolved", MISSING
        slots, _ = _cell_diffs(fields, [base], [], [theirs])
        return "conflict", slots[0]
    if theirs is MISSING:
        if ours == base:
            return "resolved", MISSING
        slots, _ = _cell_diffs(fields, [base], [ours], [])
        return "conflict", slots[0]
    if ours == theirs:
        return "resolved", ours
    if ours == base:
        return "resolved", theirs
    if theirs == base:
        return "resolved", ours

    resolved = []
    has_conflict = False
    for index in range(len(fields)):
        base_value = base[index]
        ours_value = ours[index]
        theirs_value = theirs[index]
        if ours_value == theirs_value:
            resolved.append(ours_value)
        elif ours_value == base_value:
            resolved.append(theirs_value)
        elif theirs_value == base_value:
            resolved.append(ours_value)
        else:
            has_conflict = True
            resolved.append(theirs_value)
    if not has_conflict:
        return "resolved", resolved
    slots, _ = _cell_diffs(fields, [base], [ours], [theirs])
    return "conflict", slots[0]


def _resolve_conflict_block(
    fields: tuple[str, ...],
    base_rows: list[list[Any]],
    ours_rows: list[list[Any]],
    theirs_rows: list[list[Any]],
) -> list[dict[str, Any]]:
    if not base_rows:
        return [{
            "type": "resolved",
            "policy": "same_anchor_insert_union_ours_first",
            "ours": ours_rows,
            "base": [],
            "theirs": theirs_rows,
            "rows": _merge_inserted_rows(ours_rows, theirs_rows),
        }]

    ours_states, ours_insertions = _row_edit_plan(base_rows, ours_rows)
    theirs_states, theirs_insertions = _row_edit_plan(base_rows, theirs_rows)
    segments: list[dict[str, Any]] = []
    resolved_rows: list[list[Any]] = []

    def flush_resolved() -> None:
        nonlocal resolved_rows
        if resolved_rows:
            segments.append({
                "type": "resolved",
                "policy": "continuous_anchor_three_way",
                "rows": resolved_rows,
            })
            resolved_rows = []

    for anchor in range(len(base_rows) + 1):
        resolved_rows.extend(_merge_inserted_rows(
            ours_insertions.get(anchor, []), theirs_insertions.get(anchor, [])
        ))
        if anchor == len(base_rows):
            break
        status, value = _merge_base_row(
            fields, base_rows[anchor], ours_states[anchor], theirs_states[anchor]
        )
        if status == "resolved":
            if value is not MISSING:
                resolved_rows.append(value)
            continue
        flush_resolved()
        ours = ours_states[anchor]
        theirs = theirs_states[anchor]
        segments.append({
            "type": "conflict",
            "ours": [] if ours is MISSING else [ours],
            "base": [base_rows[anchor]],
            "theirs": [] if theirs is MISSING else [theirs],
            "alignmentRequired": ours is MISSING or theirs is MISSING,
            "rowSlots": [value],
        })
    flush_resolved()
    return segments


def _parse_merge(output: str, fields: tuple[str, ...]) -> list[dict[str, Any]]:
    segments: list[dict[str, Any]] = []
    normal: list[str] = []
    ours: list[str] = []
    base: list[str] = []
    theirs: list[str] = []
    state = "normal"

    def flush_normal() -> None:
        nonlocal normal
        if normal:
            segments.append({"type": "merged", "rows": _decode_rows(normal)})
            normal = []

    for line in output.splitlines():
        if state == "normal" and line.startswith("<<<<<<< OURS"):
            flush_normal()
            state = "ours"
            ours, base, theirs = [], [], []
        elif state == "ours" and line.startswith("||||||| BASE"):
            state = "base"
        elif state == "base" and line == "=======":
            state = "theirs"
        elif state == "theirs" and line.startswith(">>>>>>> THEIRS"):
            decoded_ours = _decode_rows(ours)
            decoded_base = _decode_rows(base)
            decoded_theirs = _decode_rows(theirs)
            segments.extend(_resolve_conflict_block(
                fields, decoded_base, decoded_ours, decoded_theirs
            ))
            state = "normal"
        elif state == "normal":
            normal.append(line)
        elif state == "ours":
            ours.append(line)
        elif state == "base":
            base.append(line)
        else:
            theirs.append(line)

    if state != "normal":
        raise BridgeError(f"Unterminated git conflict marker block ({state}).")
    flush_normal()
    return segments


def _sheet_lines(ws) -> tuple[str, ...]:
    lines = []
    for row in range(1, (ws.max_row or 0) + 1):
        values = [_json_value(ws.cell(row, column).value) for column in range(1, (ws.max_column or 0) + 1)]
        while values and values[-1] is None:
            values.pop()
        if values and any(value is not None for value in values):
            lines.append(_line(values))
    return tuple(lines)


def _sheet_similarity(left, right) -> float:
    return _column_similarity(_sheet_lines(left), _sheet_lines(right))


def _match_base_sheets(base_workbook, side_workbook) -> tuple[dict[str, str], set[str]]:
    matches = {
        name: name for name in base_workbook.sheetnames if name in side_workbook.sheetnames
    }
    remaining_base = [name for name in base_workbook.sheetnames if name not in matches]
    remaining_side = [name for name in side_workbook.sheetnames if name not in matches.values()]
    scores = {
        (base_name, side_name): _sheet_similarity(
            base_workbook[base_name], side_workbook[side_name]
        )
        for base_name in remaining_base
        for side_name in remaining_side
    }
    while remaining_base and remaining_side:
        candidates = []
        for base_name in remaining_base:
            ranked = sorted(
                ((scores[(base_name, side_name)], side_name) for side_name in remaining_side),
                reverse=True,
            )
            if not ranked or ranked[0][0] < 0.60:
                continue
            if len(ranked) > 1 and ranked[0][0] == ranked[1][0]:
                continue
            score, side_name = ranked[0]
            reverse_ranked = sorted(
                ((scores[(candidate, side_name)], candidate) for candidate in remaining_base),
                reverse=True,
            )
            if len(reverse_ranked) > 1 and reverse_ranked[0][0] == reverse_ranked[1][0]:
                continue
            if reverse_ranked[0][1] == base_name:
                candidates.append((score, base_name, side_name))
        if not candidates:
            break
        _, base_name, side_name = max(candidates)
        matches[base_name] = side_name
        remaining_base.remove(base_name)
        remaining_side.remove(side_name)
    return matches, set(remaining_side)


def _sheet_triplets(base_workbook, ours_workbook, theirs_workbook) -> list[dict[str, str | None]]:
    ours_matches, ours_only = _match_base_sheets(base_workbook, ours_workbook)
    theirs_matches, theirs_only = _match_base_sheets(base_workbook, theirs_workbook)
    triplets = [
        {
            "logical": base_name,
            "base": base_name,
            "ours": ours_matches.get(base_name),
            "theirs": theirs_matches.get(base_name),
        }
        for base_name in base_workbook.sheetnames
    ]
    for name in sorted(ours_only & theirs_only):
        triplets.append({"logical": name, "base": None, "ours": name, "theirs": name})
    for name in sorted(theirs_only - ours_only):
        triplets.append({"logical": name, "base": None, "ours": None, "theirs": name})
    for name in sorted(ours_only - theirs_only):
        triplets.append({"logical": name, "base": None, "ours": name, "theirs": None})
    return triplets


def _sheet_rename_status(base_name: str | None, ours_name: str | None, theirs_name: str | None) -> str:
    if base_name is None:
        return "not_applicable"
    ours_changed = ours_name is not None and ours_name != base_name
    theirs_changed = theirs_name is not None and theirs_name != base_name
    if ours_changed and theirs_changed:
        return "renamed_both_same" if ours_name == theirs_name else "rename_conflict"
    if ours_changed:
        return "renamed_ours"
    if theirs_changed:
        return "renamed_theirs"
    return "unchanged"


def _presence_action(base_exists: bool, ours_exists: bool, theirs_exists: bool) -> str:
    if not base_exists:
        if ours_exists and theirs_exists:
            return "added_both"
        return "added_ours" if ours_exists else "added_theirs"
    if not ours_exists and not theirs_exists:
        return "deleted_both"
    if not ours_exists:
        return "deleted_ours_or_modified_theirs"
    if not theirs_exists:
        return "deleted_theirs_or_modified_ours"
    return "merge"


def analyze_workbooks(
    base_path: Path | str,
    ours_path: Path | str,
    theirs_path: Path | str,
    output_dir: Path | str,
    *,
    ignore_columns: Iterable[str] = (),
    write_sidecars: bool = True,
    skip_hidden_sheets: bool = True,
) -> dict[str, Any]:
    base_path = Path(base_path).resolve()
    ours_path = Path(ours_path).resolve()
    theirs_path = Path(theirs_path).resolve()
    output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    ignore_columns = tuple(str(name) for name in ignore_columns)
    ignored = {str(name).casefold() for name in ignore_columns}

    wb_base = load_workbook(base_path, data_only=False, keep_links=False)
    wb_ours = load_workbook(ours_path, data_only=False, keep_links=False)
    wb_theirs = load_workbook(theirs_path, data_only=False, keep_links=False)
    try:
        sheets = []
        total_conflicts = 0
        total_column_conflicts = 0
        total_sheet_conflicts = 0
        clean_merges = 0
        triplets = _sheet_triplets(wb_base, wb_ours, wb_theirs)
        for sheet_index, triplet in enumerate(triplets, 1):
            name = str(triplet["logical"])
            base_name = triplet["base"]
            ours_name = triplet["ours"]
            theirs_name = triplet["theirs"]
            if skip_hidden_sheets:
                # 隐藏页签（工具派生表）默认不参与冲突检测：以本地可见性为准
                hidden = False
                if ours_name is not None and ours_name in wb_ours.sheetnames:
                    hidden = wb_ours[ours_name].sheet_state == "hidden"
                elif theirs_name is not None and theirs_name in wb_theirs.sheetnames:
                    hidden = wb_theirs[theirs_name].sheet_state == "hidden"
                if hidden:
                    continue
            base_exists = base_name is not None
            ours_exists = ours_name is not None
            theirs_exists = theirs_name is not None
            action = _presence_action(base_exists, ours_exists, theirs_exists)
            rename_status = _sheet_rename_status(base_name, ours_name, theirs_name)
            sheet_report: dict[str, Any] = {
                "name": name,
                "names": {"base": base_name, "ours": ours_name, "theirs": theirs_name},
                "action": action,
                "renameStatus": rename_status,
                "present": {"base": base_exists, "ours": ours_exists, "theirs": theirs_exists},
            }
            if action != "merge":
                sheet_conflict = None
                if not base_exists and ours_exists and theirs_exists:
                    sheet_report.update({
                        "status": "clean",
                        "resolution": "theirs",
                        "policy": "no_base_remote_wins",
                    })
                elif not base_exists:
                    sheet_report.update({
                        "status": "clean",
                        "resolution": "ours" if ours_exists else "theirs",
                    })
                elif not ours_exists and not theirs_exists:
                    sheet_report.update({"status": "clean", "resolution": "deleted_both"})
                else:
                    survivor = "ours" if ours_exists else "theirs"
                    survivor_name = ours_name if ours_exists else theirs_name
                    survivor_workbook = wb_ours if ours_exists else wb_theirs
                    unchanged = (
                        survivor_name == base_name
                        and _sheet_lines(wb_base[base_name]) == _sheet_lines(survivor_workbook[survivor_name])
                    )
                    if unchanged:
                        sheet_report.update({
                            "status": "clean",
                            "resolution": "theirs" if not ours_exists else "ours",
                        })
                    else:
                        deleted = "ours" if not ours_exists else "theirs"
                        sheet_conflict = {
                            "kind": f"delete_modify_{deleted}",
                            "survivor": survivor,
                            "survivorName": survivor_name,
                        }
                        sheet_report.update({"status": "conflict", "sheetConflict": sheet_conflict})
                        total_sheet_conflicts += 1
                sheets.append(sheet_report)
                continue

            layouts = {
                "base": _layout(wb_base[base_name]),
                "ours": _layout(wb_ours[ours_name]),
                "theirs": _layout(wb_theirs[theirs_name]),
            }
            worksheets = {
                "base": wb_base[base_name],
                "ours": wb_ours[ours_name],
                "theirs": wb_theirs[theirs_name],
            }
            layouts, column_alignment, column_alignment_mode = _align_layouts(
                worksheets, layouts
            )
            common = set(layouts["base"].fields) & set(layouts["ours"].fields) & set(layouts["theirs"].fields)
            projection_fields = tuple(
                field for field in layouts["base"].fields
                if field in common
                and field.casefold() not in ignored
                and not (layouts["base"].var_row is not None and field == "@1")
            )
            if not projection_fields:
                sheet_report.update({
                    "status": "unsupported",
                    "error": "No common projection fields remain after column alignment/ignore filtering.",
                })
                sheets.append(sheet_report)
                continue

            rows = {
                "base": _rows(wb_base[base_name], layouts["base"], projection_fields),
                "ours": _rows(wb_ours[ours_name], layouts["ours"], projection_fields),
                "theirs": _rows(wb_theirs[theirs_name], layouts["theirs"], projection_fields),
            }
            row_alignment = _row_alignment_report(rows)
            column_conflict_count = _annotate_column_conflicts(column_alignment, rows)
            total_column_conflicts += column_conflict_count
            sheet_dir = output_dir / _safe_name(name, sheet_index)
            sheet_dir.mkdir(parents=True, exist_ok=True)
            paths = {}
            for side in ("base", "ours", "theirs"):
                projection_path = sheet_dir / f"{side}.jsonl"
                _write_projection(projection_path, rows[side])
                if write_sidecars:
                    sidecar_path = sheet_dir / f"{side}.sidecar.json"
                    _write_sidecar(sidecar_path, layouts[side], rows[side])
                paths[side] = projection_path

            merged_text, raw_git_clean, git_stderr = _git_merge_file(
                paths["ours"], paths["base"], paths["theirs"]
            )
            merge_output_path = sheet_dir / "git-merge-output.txt"
            _write_text_lf(merge_output_path, merged_text)
            segments = _parse_merge(merged_text, projection_fields)
            conflict_count = sum(segment["type"] == "conflict" for segment in segments)
            total_conflicts += conflict_count
            git_clean = conflict_count == 0
            sheet_conflict_count = int(rename_status == "rename_conflict")
            total_sheet_conflicts += sheet_conflict_count
            sheet_clean = git_clean and column_conflict_count == 0 and sheet_conflict_count == 0
            clean_merges += int(sheet_clean)
            all_fields = [entry["field"] for entry in column_alignment]
            sheet_report.update({
                "status": "clean" if sheet_clean else "conflict",
                "gitStatus": "clean" if git_clean else "conflict",
                "gitRawStatus": "clean" if raw_git_clean else "conflict",
                "columnAlignmentMode": column_alignment_mode,
                "columnAlignment": column_alignment,
                "columnConflictCount": column_conflict_count,
                "rowAlignment": row_alignment,
                "sheetConflictCount": sheet_conflict_count,
                "projectionFields": list(projection_fields),
                "ignoredFields": [field for field in layouts["base"].fields if field.casefold() in ignored],
                "columns": {
                    "base": list(layouts["base"].fields),
                    "ours": list(layouts["ours"].fields),
                    "theirs": list(layouts["theirs"].fields),
                    "baseOnly": sorted(set(layouts["base"].fields) - set(layouts["ours"].fields) - set(layouts["theirs"].fields)),
                    "oursOnly": sorted(set(layouts["ours"].fields) - set(layouts["base"].fields)),
                    "theirsOnly": sorted(set(layouts["theirs"].fields) - set(layouts["base"].fields)),
                    "all": all_fields,
                },
                "rowCounts": {side: len(rows[side]) for side in rows},
                "conflictBlockCount": conflict_count,
                "segments": segments,
                "artifacts": {
                    "directory": str(sheet_dir),
                    "gitMergeOutput": str(merge_output_path),
                    "base": str(paths["base"]),
                    "ours": str(paths["ours"]),
                    "theirs": str(paths["theirs"]),
                },
                "gitStderr": git_stderr,
            })
            sheets.append(sheet_report)

        report = {
            "version": 4,
            "inputs": {"base": str(base_path), "ours": str(ours_path), "theirs": str(theirs_path)},
            "ignoreColumns": sorted(ignore_columns),
            "summary": {
                "sheetCount": len(sheets),
                "cleanMergeSheetCount": clean_merges,
                "conflictBlockCount": total_conflicts,
                "columnConflictCount": total_column_conflicts,
                "sheetConflictCount": total_sheet_conflicts,
            },
            "sheets": sheets,
        }
        report_path = output_dir / "report.json"
        report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        report["reportPath"] = str(report_path)
        return report
    finally:
        wb_base.close()
        wb_ours.close()
        wb_theirs.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base", required=True, help="Common-ancestor XLSX path.")
    parser.add_argument("--ours", required=True, help="Local/ours XLSX path.")
    parser.add_argument("--theirs", required=True, help="Remote/theirs XLSX path.")
    parser.add_argument("--output-dir", required=True, help="Directory for JSONL, sidecars, and report.json.")
    parser.add_argument(
        "--ignore-column", action="append", default=[],
        help="Field name excluded from row-alignment JSONL; repeat for multiple fields.",
    )
    parser.add_argument("--print-report", action="store_true", help="Print the complete report instead of the summary.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        report = analyze_workbooks(
            args.base,
            args.ours,
            args.theirs,
            args.output_dir,
            ignore_columns=args.ignore_column,
        )
        output = report if args.print_report else {
            "ok": True,
            "report": report["reportPath"],
            "summary": report["summary"],
        }
        print(json.dumps(output, ensure_ascii=False, indent=2))
        return 0
    except (BridgeError, OSError, ValueError) as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False, indent=2))
        return 1


if __name__ == "__main__":
    sys.exit(main())
