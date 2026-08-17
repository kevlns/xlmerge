# -*- coding: utf-8 -*-
"""E2E：构造 xlsx 冲突仓库并驱动 xlmerge 完整流程。"""
import json
import os
import subprocess
import sys
import tempfile
import time
import urllib.request
from pathlib import Path

from openpyxl import Workbook, load_workbook

BIN = [sys.argv[1]] if not sys.argv[1].endswith(".js") else [sys.argv[2] or "node", sys.argv[1]]
REPO = Path(tempfile.mkdtemp(prefix="xlsx-e2e-")) / "meta"
REPO.mkdir(parents=True)


def git(*args, check=True):
    r = subprocess.run(["git", "-C", str(REPO), *args], capture_output=True, text=True, encoding="utf-8")
    if check and r.returncode:
        raise SystemExit(f"git {args} failed: {r.stdout}{r.stderr}")
    return r


def write(path, version):
    wb = Workbook(); ws = wb.active; ws.title = "Items"
    ws.append(["##var", "id", "name", "price"])
    ws.append(["##type", "int", "string", "float"])
    ws.append(["##", None, None, None])
    ws.append([None, 1, {"base": "sword", "ours": "sword-v2", "theirs": "sword-v3"}[version], 10.0])
    ws.append([None, 2, "shield", {"base": 20.0, "ours": 25.0, "theirs": 30.0}[version]])
    wb.save(path)


def run_cli(*args, timeout=60):
    r = subprocess.run([*BIN, "--repo", str(REPO), *args], capture_output=True, text=True, encoding="utf-8", timeout=timeout)
    return r


def api(url, path, data=None, method=None):
    """用 curl 访问 API，规避 urllib 在 Windows 上对 wsgiref 连接关闭的处理差异。"""
    cmd = ["curl", "-s", "--max-time", "10"]
    if data is not None:
        cmd += ["-X", (method or "POST"), "-H", "Content-Type: application/json", "--data-binary", "@-"]
    cmd.append(url + path)
    r = subprocess.run(cmd, input=data if isinstance(data, bytes) else (json.dumps(data).encode() if data else None), capture_output=True, timeout=20)
    return json.loads(r.stdout.decode("utf-8"))


target = REPO / "new_meta" / "Items.xlsx"
target.parent.mkdir(parents=True)

git("init", "-q", "-b", "main")
git("config", "user.email", "e2e@t"); git("config", "user.name", "e2e")
write(target, "base"); git("add", "-A"); git("commit", "-qm", "base")

git("checkout", "-q", "-b", "ours"); write(target, "ours"); git("add", "-A"); git("commit", "-qm", "ours")
git("checkout", "-q", "main"); git("checkout", "-q", "-b", "theirs"); write(target, "theirs"); git("add", "-A"); git("commit", "-qm", "theirs")
git("checkout", "-q", "main")
git("checkout", "-q", "ours")
r = git("merge", "theirs", check=False)
assert r.returncode != 0, "应产生冲突"
assert "new_meta/Items.xlsx" in git("diff", "--name-only", "--diff-filter=U").stdout, "xlsx 应处于未合并状态"

# 1) detect
r = run_cli("detect")
detect = json.loads(r.stdout)
print("detect:", json.dumps(detect, ensure_ascii=False))
assert detect["ok"] and detect["count"] == 1

# 2) launch（后台，不开浏览器）
r = run_cli("launch", "--no-browser", timeout=90)
launch = json.loads(r.stdout)
print("launch ok:", launch["ok"], "url:", launch["url"])
assert launch["ok"] and launch["url"].startswith("http://127.0.0.1:")

# 3) API 获取 diff
diff = api(launch["url"], "/api/diff")
print("diff summary:", json.dumps(diff["summary"], ensure_ascii=False))
sheets = [s["name"] for s in diff["files"][0]["diff"]["sheets"]]
assert "Items" in sheets

# 4) commit 预览 + 非交互写回（theirs 全量）
decisions = {"files": {"new_meta/Items.xlsx": {"sheets": {"Items": {"cells": {}}}}}}
# 找出冲突 Cell 并选 theirs
for sheet in diff["files"][0]["diff"]["sheets"]:
    if sheet["name"] != "Items":
        continue
    for cell in sheet.get("cells", []):
        if cell.get("resolution") == "conflict":
            decisions["files"]["new_meta/Items.xlsx"]["sheets"]["Items"]["cells"][cell["id"]] = "theirs"
print("cells picked:", list(decisions["files"]["new_meta/Items.xlsx"]["sheets"]["Items"]["cells"]))
preview = api(launch["url"], "/api/commit-preview", json.dumps({"decisions": decisions}).encode())
print("commit-preview:", preview["message"][:120].replace("\n", " | "))

resolved = api(launch["url"], "/api/resolve", json.dumps({"decisions": decisions, "commit": True}).encode())
print("resolve success:", resolved.get("success"), resolved.get("error", ""))
assert resolved["success"]

api(launch["url"], "/api/shutdown", b"", "POST")
for _ in range(50):
    if subprocess.run(["taskkill", "/PID", str(launch["pid"]), "/F"], capture_output=True).returncode == 0:
        break
    time.sleep(0.2)

# 5) 复核仓库状态
assert git("diff", "--name-only", "--diff-filter=U").stdout.strip() == "", "不应再有未合并文件"
log = git("log", "-1", "--oneline").stdout.strip()
print("last commit:", log)

wb = load_workbook(target)
ws = wb["Items"]
assert ws["C4"].value == "sword-v3" and ws["D5"].value == 30.0, "应写入 theirs 值"
assert ws["A1"].value == "##var" and ws["A2"].value == "##type", "Luban 表头保留"
wb.close()
print("E2E PASS")
